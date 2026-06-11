#!/usr/bin/env python3
"""
=============================================================================
  Identity Management App — Active Directory Integration
  Author : Built with Python + tkinter + ldap3
  Purpose: Menu-driven AD User Account & Group Management
  Requires: Python 3.8+, tkinter, ldap3
  Install : pip install ldap3
  Run     : python3 identity_manager.py
=============================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import ldap3
from ldap3 import Server, Connection, ALL, MODIFY_REPLACE, MODIFY_ADD, Tls
from ldap3.core.exceptions import LDAPException
import ssl as _ssl
import threading
import json
import datetime
import re
import os
import hashlib
import base64
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT / STATS  (in-memory + persisted to config)
# ─────────────────────────────────────────────────────────────────────────────
_STATS_FILE = Path.home() / ".idmanager_stats.json"

def _load_stats():
    try:
        if _STATS_FILE.exists():
            return json.loads(_STATS_FILE.read_text())
    except Exception:
        pass
    return {"users_created": 0, "users_deleted": 0, "users_modified": 0,
            "groups_created": 0, "groups_deleted": 0, "groups_modified": 0,
            "audit_log": []}

def _save_stats(s: dict):
    try:
        s["audit_log"] = s.get("audit_log", [])[-500:]
        _STATS_FILE.write_text(json.dumps(s, indent=2))
    except Exception:
        pass

def _record(action: str, object_type: str, name: str):
    """Record an auditable event and increment the relevant counter."""
    s = _load_stats()
    # action values: create, modif, delet  →  key: users_created, users_modified, users_deleted
    suffix_map = {"create": "created", "modif": "modified", "delet": "deleted"}
    key = f"{object_type}s_{suffix_map.get(action, action)}"
    s[key] = s.get(key, 0) + 1
    s.setdefault("audit_log", []).append({
        "ts":     datetime.datetime.now().isoformat(timespec="seconds"),
        "action": suffix_map.get(action, action),
        "type":   object_type,
        "name":   name,
    })
    _save_stats(s)


# ─────────────────────────────────────────────────────────────────────────────
#  THEME CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
NAVY      = "#0F2D5E"
ACCENT    = "#1D6FD8"
ACCENT2   = "#1558B0"
TEAL      = "#0D7A6B"
GREEN     = "#166534"
RED       = "#991B1B"
AMBER     = "#92400E"
LIGHT_BLU = "#EBF3FF"
LIGHT_GRN = "#DCFCE7"
LIGHT_RED = "#FEE2E2"
BG        = "#F0F4F8"
SURFACE   = "#FFFFFF"
SURFACE2  = "#F8FAFC"
BORDER    = "#CBD5E1"
TEXT      = "#1E293B"
TEXT2     = "#475569"
TEXT3     = "#94A3B8"
WHITE     = "#FFFFFF"
FONT_HEAD = ("Arial", 14, "bold")
FONT_SUB  = ("Arial", 11, "bold")
FONT_BODY = ("Arial", 10)
FONT_MONO = ("Courier New", 9)
FONT_TINY = ("Arial", 8)

# ─────────────────────────────────────────────────────────────────────────────
#  SCROLL HELPER — mouse-wheel / trackpad works on whichever canvas is hovered
# ─────────────────────────────────────────────────────────────────────────────
def _bind_mousewheel(canvas):
    """
    Attach mouse-wheel scroll to *canvas*.
    Uses bind_all while the pointer is inside the canvas so the scroll 
    reaches the canvas even when hovering over child widgets (labels, entries).
    On leave the bind_all handler is removed so other canvases are not affected.
    """
    def _on_wheel(event):
        if event.delta:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_btn4(event):
        canvas.yview_scroll(-1, "units")

    def _on_btn5(event):
        canvas.yview_scroll(1, "units")

    def _enter(event):
        canvas.bind_all("<MouseWheel>", _on_wheel)
        canvas.bind_all("<Button-4>",   _on_btn4)
        canvas.bind_all("<Button-5>",   _on_btn5)

    def _leave(event):
        canvas.unbind_all("<MouseWheel>")
        canvas.unbind_all("<Button-4>")
        canvas.unbind_all("<Button-5>")

    canvas.bind("<Enter>", _enter, add="+")
    canvas.bind("<Leave>", _leave, add="+")



# ─────────────────────────────────────────────────────────────────────────────
#  PERSISTENT CONFIG + ACTIVATION
# ─────────────────────────────────────────────────────────────────────────────
CONFIG_FILE  = Path.home() / ".idmanager_config.json"
LOGO_FILE    = Path.home() / ".idmanager_logo.dat"
# Activation: each code = HMAC of "IDManager|DAYS|SALT" with secret key
_ACT_KEY     = "ITProAcademy2024SecretKey"
_VENDOR_NAME = "ITProAcademy.co.in"
_APP_VERSION = "1.0"

def _load_config():
    """Load saved config from disk."""
    try:
        if CONFIG_FILE.exists():
            return json.loads(CONFIG_FILE.read_text())
    except Exception:
        pass
    return {}

def _save_config(data: dict):
    """Save config to disk."""
    try:
        CONFIG_FILE.write_text(json.dumps(data, indent=2))
    except Exception:
        pass

def _load_logo_data():
    """Return base64 logo bytes if saved, else None."""
    try:
        if LOGO_FILE.exists():
            return LOGO_FILE.read_bytes()
    except Exception:
        pass
    return None

def _save_logo_data(raw_bytes):
    try:
        LOGO_FILE.write_bytes(raw_bytes)
    except Exception:
        pass

def _random_salt(length: int = 8) -> str:
    """Generate a cryptographically random uppercase alphanumeric salt."""
    import string as _string
    alphabet = _string.ascii_uppercase + _string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))

def _generate_activation_code(days: int, salt: str = None) -> str:
    """
    Generate a unique activation code for N days.
    Format: IDP-{DAYS:04d}-{HASH16}-{SALT8}
    Every call produces a different code because a fresh random salt is used.
    """
    if salt is None:
        salt = _random_salt(8)
    msg = f"IDManager|{days}|{_ACT_KEY}|{salt}"
    h   = hashlib.sha256(msg.encode()).hexdigest()[:16].upper()
    return f"IDP-{days:04d}-{h}-{salt}"

def _verify_activation_code(code: str):
    """
    Returns number of days if valid, else 0.
    Supports both new 4-part format (IDP-DAYS-HASH-SALT)
    and legacy 3-part format (IDP-DAYS-HASH) for backward compatibility.
    """
    try:
        parts = code.strip().upper().split("-")
        if len(parts) < 3 or parts[0] != "IDP":
            return 0
        days = int(parts[1])
        if days < 1:
            return 0

        if len(parts) == 4:
            # New format: IDP-DAYS-HASH16-SALT8
            h_part = parts[2]
            salt   = parts[3]
            msg    = f"IDManager|{days}|{_ACT_KEY}|{salt}"
            expected = hashlib.sha256(msg.encode()).hexdigest()[:16].upper()
            return days if h_part == expected else 0

        elif len(parts) == 3:
            # Legacy format: IDP-DAYS-HASH12 (fixed key, no salt)
            h_part   = parts[2]
            msg      = f"IDManager|{days}|{_ACT_KEY}"
            expected = hashlib.sha256(msg.encode()).hexdigest()[:12].upper()
            return days if h_part == expected else 0

    except Exception:
        pass
    return 0

def _get_activation_status(cfg: dict):
    """Returns (is_active, days_remaining, expiry_str)."""
    activated_on = cfg.get("activated_on")
    total_days   = cfg.get("activated_days", 0)
    if not activated_on or not total_days:
        return False, 0, "Not activated"
    try:
        from datetime import date
        start   = datetime.date.fromisoformat(activated_on)
        end     = start + datetime.timedelta(days=total_days)
        today   = datetime.date.today()
        rem     = (end - today).days
        if rem > 0:
            return True, rem, end.isoformat()
        return False, 0, f"Expired {end.isoformat()}"
    except Exception:
        return False, 0, "Unknown"

# ─────────────────────────────────────────────────────────────────────────────
#  AD CONNECTOR  (wraps ldap3)
# ─────────────────────────────────────────────────────────────────────────────
class ADConnector:
    def __init__(self):
        self.conn      = None
        self.server    = None
        self.base_dn   = ""
        self.connected = False
        self.use_ssl   = False
        # stored for StartTLS password connections
        self._host     = ""
        self._port     = 389
        self._user     = ""
        self._pwd      = ""

    def connect(self, host, port, base_dn, username, password, use_ssl=False):
        """
        Connect to AD. If SSL is requested but fails, automatically retries
        without SSL on port 389 so the user is never left stranded by a
        bad saved SSL setting.
        """
        port = int(port)

        def _try(h, p, ssl):
            try:
                srv  = Server(h, port=p, use_ssl=ssl, get_info=ALL, connect_timeout=10)
                conn = Connection(srv, user=username, password=password, auto_bind=True)
                return True, srv, conn, ssl, ""
            except Exception as e:
                return False, None, None, ssl, str(e)

        ok, srv, conn, actual_ssl, err = _try(host, port, use_ssl)

        # Auto-retry: if SSL requested but failed, retry without SSL on 389
        if not ok and use_ssl:
            retry_port = 389 if port == 636 else port
            ok2, srv2, conn2, _, err2 = _try(host, retry_port, False)
            if ok2:
                ok, srv, conn, actual_ssl, err = ok2, srv2, conn2, False, ""
                # err left as "" — success message will note the fallback
                fallback_note = (f" (Note: LDAPS failed, connected via plain LDAP on port "
                                 f"{retry_port}. Passwords will use StartTLS.)")
            else:
                # Both failed — report the original SSL error, not the retry
                pass
        else:
            fallback_note = ""

        if ok:
            self.server    = srv
            self.conn      = conn
            self.base_dn   = base_dn
            self.use_ssl   = actual_ssl
            self.connected = True
            self._host     = host
            self._port     = self.conn.server.port
            self._user     = username
            self._pwd      = password
            mode = "LDAPS (SSL)" if actual_ssl else "LDAP (plain)"
            return True, f"Connected successfully via {mode}.{fallback_note}"
        else:
            self.connected = False
            # Parse the error for a friendlier message
            err_lower = err.lower()
            if "10054" in err or "forcibly closed" in err_lower or "wrapping" in err_lower:
                friendly = (f"SSL/TLS handshake failed — the server closed the connection. "
                            f"This usually means port {port} is not configured for LDAPS on this DC. "
                            f"Try unchecking 'Use SSL' to connect on plain LDAP (port 389); "
                            f"passwords will be set via StartTLS automatically.")
            elif "invalidcredentials" in err_lower or "error 49" in err_lower:
                friendly = "Invalid credentials — check username (UPN format: user@domain.com) and password."
            elif "timeout" in err_lower or "timed out" in err_lower:
                friendly = f"Connection timed out — cannot reach {host}:{port}. Check firewall and hostname."
            elif "name or service not known" in err_lower or "getaddrinfo" in err_lower:
                friendly = f"Cannot resolve hostname '{host}'. Check the AD Server field."
            else:
                friendly = err
            return False, friendly

    def set_password(self, user_dn: str, new_password: str) -> tuple:
        """
        Try every available method to set an AD password.
        Returns (success: bool, method_used: str, last_error: str).

        Method order (least to most complex):
          1. unicodePwd on current connection  — instant if already LDAPS
          2. Extended Password Modify op        — works on some DCs over plain LDAP
          3. Dedicated LDAPS connection (636)  — opens a fresh SSL conn just for pwd
          4. StartTLS on port 389              — TLS upgrade on existing port
        """
        encoded_pwd = ('"' + new_password + '"').encode("utf-16-le")
        last_err    = ""

        # ── Method 1: unicodePwd on existing connection ──────────────────────
        # Works immediately when the main connection is LDAPS (use_ssl=True).
        try:
            ok = self.conn.modify(user_dn, {"unicodePwd": [(MODIFY_REPLACE, [encoded_pwd])]})
            if ok:
                return True, "unicodePwd (existing connection)", ""
            last_err = (f"{self.conn.result.get('description','')} — "
                        f"{self.conn.result.get('message','')}")
        except Exception as e:
            last_err = str(e)

        # ── Method 2: ldap3 extended Password Modify op ──────────────────────
        # Works over plain LDAP on DCs that allow it (not all do).
        try:
            ok = self.conn.extend.microsoft.modify_password(user_dn, new_password)
            if ok:
                return True, "Extended Password op", ""
            last_err = str(self.conn.result)
        except Exception as e:
            last_err = str(e)

        # ── Method 3: dedicated LDAPS connection on port 636 ─────────────────
        # Even when the main session is plain LDAP (389), we open a SEPARATE
        # SSL connection on 636 purely for the unicodePwd write, then close it.
        # This is the most reliable path for DCs with LDAPServerIntegrity=Require
        # and StartTLS disabled.
        try:
            tls_cfg = Tls(validate=_ssl.CERT_NONE)
            tls_cfg._ctx = _ssl.SSLContext(_ssl.PROTOCOL_TLS_CLIENT)
            tls_cfg._ctx.check_hostname = False
            tls_cfg._ctx.verify_mode    = _ssl.CERT_NONE

            ssl_srv  = Server(self._host, port=636, use_ssl=True,
                              tls=tls_cfg, get_info=ALL, connect_timeout=10)
            ssl_conn = Connection(ssl_srv, user=self._user, password=self._pwd,
                                  auto_bind=True)
            ok = ssl_conn.modify(user_dn, {"unicodePwd": [(MODIFY_REPLACE, [encoded_pwd])]})
            if ok:
                ssl_conn.unbind()
                return True, "unicodePwd via dedicated LDAPS (port 636)", ""
            last_err = (f"{ssl_conn.result.get('description','')} — "
                        f"{ssl_conn.result.get('message','')}")
            ssl_conn.unbind()
        except Exception as e:
            last_err = str(e)

        # ── Method 4: StartTLS upgrade on port 389 ───────────────────────────
        # Opens a new connection on port 389 and upgrades to TLS before binding.
        # Skipped if DC advertises StartTLS as unavailable, but worth trying.
        try:
            tls_cfg2 = Tls(validate=_ssl.CERT_NONE)
            tls_cfg2._ctx = _ssl.SSLContext(_ssl.PROTOCOL_TLS_CLIENT)
            tls_cfg2._ctx.check_hostname = False
            tls_cfg2._ctx.verify_mode    = _ssl.CERT_NONE

            stls_srv  = Server(self._host, port=self._port, use_ssl=False,
                               tls=tls_cfg2, get_info=ALL, connect_timeout=10)
            stls_conn = Connection(stls_srv, user=self._user, password=self._pwd,
                                   auto_bind=False)
            stls_conn.open()
            if stls_conn.start_tls():
                stls_conn.bind()
                ok = stls_conn.modify(user_dn,
                    {"unicodePwd": [(MODIFY_REPLACE, [encoded_pwd])]})
                if ok:
                    stls_conn.unbind()
                    return True, "unicodePwd via StartTLS (port 389)", ""
                last_err = (f"{stls_conn.result.get('description','')} — "
                            f"{stls_conn.result.get('message','')}")
            else:
                last_err = "StartTLS unavailable on this DC"
            stls_conn.unbind()
        except Exception as e:
            last_err = str(e)

        return False, "", last_err

    def disconnect(self):
        if self.conn:
            self.conn.unbind()
        self.connected = False

    # ── USER EDIT / DELETE ────────────────────────────────────────
    def get_user_attributes(self, dn: str) -> dict:
        """Fetch all editable attributes for a user DN."""
        attrs = ["displayName","givenName","sn","userPrincipalName","sAMAccountName",
                 "mail","title","company","department","physicalDeliveryOfficeName",
                 "manager","streetAddress","l","st","postalCode","co","c",
                 "telephoneNumber","facsimileTelephoneNumber","userAccountControl",
                 "lockoutTime","memberOf"]
        try:
            from ldap3 import SUBTREE, BASE
            self.conn.search(search_base=dn, search_filter="(objectClass=*)",
                             search_scope=BASE, attributes=attrs)
            if self.conn.entries:
                e = self.conn.entries[0]
                result = {}
                for a in attrs:
                    try:
                        v = getattr(e, a)
                        result[a] = str(v) if v else ""
                    except Exception:
                        result[a] = ""
                return result
        except Exception:
            pass
        return {}

    def modify_user(self, dn: str, changes: dict) -> tuple:
        """Apply a dict of {attr: new_value} to a user. Empty string = clear."""
        ldap_changes = {}
        for attr, val in changes.items():
            if val:
                ldap_changes[attr] = [(MODIFY_REPLACE, [val])]
            else:
                ldap_changes[attr] = [(MODIFY_REPLACE, [])]
        ok = self.conn.modify(dn, ldap_changes)
        if ok:
            return True, "User updated successfully."
        return False, (f"{self.conn.result.get('description','')} — "
                       f"{self.conn.result.get('message','')}")

    def delete_user(self, dn: str) -> tuple:
        ok = self.conn.delete(dn)
        if ok:
            return True, f"User deleted: {dn}"
        return False, (f"{self.conn.result.get('description','')} — "
                       f"{self.conn.result.get('message','')}")

    # ── GROUP FULL FETCH / EDIT / DELETE ─────────────────────────
    def get_all_groups_full(self):
        """Return list of dicts with full group attributes."""
        if not self.connected:
            return []
        try:
            from ldap3 import SUBTREE
            ok = self.conn.search(
                search_base=self.base_dn,
                search_filter="(objectClass=group)",
                search_scope=SUBTREE,
                attributes=["cn","distinguishedName","description","groupType",
                            "member","memberOf","sAMAccountName","mail"],
                size_limit=500
            )
            if not ok:
                return []
            results = []
            for e in self.conn.entries:
                try:
                    gtype_raw = int(str(e.groupType)) if e.groupType else -2147483646
                    gtype_map = {
                        -2147483646: "Global Security",
                        -2147483644: "Domain Local Security",
                        -2147483640: "Universal Security",
                        2: "Global Distribution",
                        4: "Domain Local Dist.",
                        8: "Universal Distribution",
                    }
                    members = e.member.values if hasattr(e.member, "values") else []
                    results.append({
                        "cn"         : str(e.cn) if e.cn else "",
                        "dn"         : str(e.distinguishedName).strip() if e.distinguishedName else "",
                        "description": str(e.description) if e.description else "",
                        "groupType"  : gtype_map.get(gtype_raw, "Global Security"),
                        "groupTypeRaw": gtype_raw,
                        "memberCount": len(members),
                        "sam"        : str(e.sAMAccountName) if e.sAMAccountName else "",
                    })
                except Exception:
                    pass
            return sorted(results, key=lambda x: x["cn"].lower())
        except Exception:
            return []

    def get_group_members(self, group_dn: str):
        """Return list of (display_name, dn) for all members of a group."""
        try:
            from ldap3 import BASE
            self.conn.search(search_base=group_dn, search_filter="(objectClass=*)",
                             search_scope=BASE, attributes=["member"])
            if not self.conn.entries:
                return []
            member_dns = self.conn.entries[0].member.values if hasattr(
                self.conn.entries[0].member, "values") else []
            results = []
            for mdn in member_dns:
                try:
                    cn = mdn.split(",")[0].replace("CN=","").replace("cn=","")
                    results.append((cn, str(mdn)))
                except Exception:
                    pass
            return results
        except Exception:
            return []

    def modify_group(self, dn: str, name: str, description: str,
                     add_member_dns: list, remove_member_dns: list) -> tuple:
        """
        Update group attributes, optionally rename the group, and manage membership.

        Renaming a group in AD requires two separate operations:
          1. conn.modify_dn()  — renames the CN (RDN) in the directory tree
          2. conn.modify()     — updates sAMAccountName to match the new CN
        A plain modify() on the 'cn' attribute alone is NOT sufficient and will
        silently succeed without changing the visible group name.
        """
        current_cn = dn.split(",")[0].replace("CN=", "").replace("cn=", "").strip()
        new_cn     = name.strip() if name else current_cn
        new_dn     = dn  # will be updated below if rename succeeds

        # ── Step 1: Rename (modify_dn) if CN changed ─────────────────────────
        if new_cn and new_cn != current_cn:
            rename_ok = self.conn.modify_dn(dn, f"CN={new_cn}")
            if not rename_ok:
                err = self.conn.result.get("description", "")
                msg = self.conn.result.get("message", "")
                return False, (f"Group rename failed [{err}]: {msg}. "
                               f"Ensure the account has 'Write' permission on the group object.")
            # Recompute the DN after rename — parent path stays the same
            parent = ",".join(dn.split(",")[1:])
            new_dn = f"CN={new_cn},{parent}"

            # Also update sAMAccountName to match the new CN (max 20 chars)
            sam_ok = self.conn.modify(new_dn,
                {"sAMAccountName": [(MODIFY_REPLACE, [new_cn[:20]])]})
            # sAMAccountName update failure is non-fatal — log but continue
            if not sam_ok:
                sam_err = self.conn.result.get("description", "")
                # We'll include this as a warning in the success message
                sam_warn = f" (sAMAccountName update warning: {sam_err})"
            else:
                sam_warn = ""

        else:
            sam_warn = ""

        # ── Step 2: Update description and other plain attributes ─────────────
        changes = {}
        if description is not None:
            changes["description"] = [(MODIFY_REPLACE, [description] if description else [])]
        if changes:
            attr_ok = self.conn.modify(new_dn, changes)
            if not attr_ok:
                return False, (f"Attribute update failed: "
                               f"{self.conn.result.get('description','')}")

        # ── Step 3: Membership changes ────────────────────────────────────────
        for mdn in add_member_dns:
            self.conn.modify(new_dn, {"member": [(MODIFY_ADD, [mdn])]})
        from ldap3 import MODIFY_DELETE
        for mdn in remove_member_dns:
            self.conn.modify(new_dn, {"member": [(MODIFY_DELETE, [mdn])]})

        return True, f"Group updated successfully.{sam_warn}"

    def delete_group(self, dn: str) -> tuple:
        ok = self.conn.delete(dn)
        if ok:
            return True, f"Group deleted: {dn}"
        return False, (f"{self.conn.result.get('description','')} — "
                       f"{self.conn.result.get('message','')}")

    # ── REPORTS DATA ──────────────────────────────────────────────
    def get_users_full(self):
        """Return rich user list for reports."""
        if not self.connected:
            return []
        try:
            from ldap3 import SUBTREE
            self.conn.search(
                search_base=self.base_dn,
                search_filter="(&(objectClass=user)(objectCategory=person))",
                search_scope=SUBTREE,
                attributes=["displayName","sAMAccountName","mail","department",
                            "title","company","telephoneNumber","userAccountControl",
                            "whenCreated","distinguishedName"],
                size_limit=2000
            )
            results = []
            for e in self.conn.entries:
                try:
                    uac = int(str(e.userAccountControl)) if e.userAccountControl else 512
                    results.append({
                        "Name"       : str(e.displayName)       if e.displayName       else "",
                        "Username"   : str(e.sAMAccountName)    if e.sAMAccountName    else "",
                        "Email"      : str(e.mail)              if e.mail              else "",
                        "Department" : str(e.department)        if e.department        else "",
                        "Title"      : str(e.title)             if e.title             else "",
                        "Company"    : str(e.company)           if e.company           else "",
                        "Phone"      : str(e.telephoneNumber)   if e.telephoneNumber   else "",
                        "Status"     : "Disabled" if (uac & 2) else "Enabled",
                        "Created"    : str(e.whenCreated)[:10]  if e.whenCreated       else "",
                        "DN"         : str(e.distinguishedName).strip() if e.distinguishedName else "",
                    })
                except Exception:
                    pass
            return results
        except Exception:
            return []

    def get_all_ous(self):
        """Return list of (display_label, full_dn) for every OU."""
        if not self.connected:
            return []
        results = []
        try:
            from ldap3 import SUBTREE
            self.conn.search(
                search_base   = self.base_dn,
                search_filter = "(objectClass=organizationalUnit)",
                search_scope  = SUBTREE,
                attributes    = ["distinguishedName", "name"],
                size_limit    = 500
            )
            for e in self.conn.entries:
                try:
                    dn = str(e.distinguishedName).strip() if e.distinguishedName else ""
                    if dn and dn.lower() != self.base_dn.lower():
                        # Build a friendly label: extract OU= components only,
                        # reverse so parent → child reads left to right, join with >
                        # e.g. "OU=IT,OU=Users,DC=..." → "Users > IT"
                        try:
                            suffix = "," + self.base_dn
                            rel    = dn[: -len(suffix)] if dn.lower().endswith(suffix.lower()) else dn
                            ou_parts = [p.strip()[3:] for p in rel.split(",")
                                        if p.strip().upper().startswith("OU=")]
                            ou_parts.reverse()          # parent first → child last
                            label = " > ".join(ou_parts) if ou_parts else rel
                        except Exception:
                            label = dn
                        results.append((label, dn))
                except Exception:
                    pass
            results.sort(key=lambda x: x[0].lower())
        except Exception as ex:
            self._last_error = str(ex)
        # Always prepend root DN as first option
        results.insert(0, (f"[Root — {self.base_dn}]", self.base_dn))
        return results

    # ── USERS ────────────────────────────────────────────────────
    def get_all_groups(self):
        if not self.connected:
            return []
        try:
            from ldap3 import SUBTREE
            ok = self.conn.search(
                search_base   = self.base_dn,
                search_filter = "(objectClass=group)",
                search_scope  = SUBTREE,
                attributes    = ["cn", "distinguishedName"],
                size_limit    = 500
            )
            if not ok:
                self._last_error = str(self.conn.result)
                return []
            results = []
            for e in self.conn.entries:
                try:
                    cn_val = str(e.cn) if e.cn else ""
                    dn_val = str(e.distinguishedName).strip() if e.distinguishedName else ""
                    if cn_val and dn_val:
                        results.append((cn_val, dn_val))
                except Exception:
                    pass
            return sorted(results, key=lambda x: x[0].lower())
        except Exception as ex:
            self._last_error = str(ex)
            return []

    def get_all_users(self):
        if not self.connected:
            return []
        try:
            from ldap3 import SUBTREE
            ok = self.conn.search(
                search_base   = self.base_dn,
                search_filter = "(&(objectClass=user)(objectCategory=person))",
                search_scope  = SUBTREE,
                attributes    = ["displayName", "sAMAccountName", "distinguishedName", "mail"],
                size_limit    = 1000
            )
            if not ok:
                self._last_error = str(self.conn.result)
                return []
            results = []
            for e in self.conn.entries:
                try:
                    dn   = str(e.distinguishedName).strip() if e.distinguishedName else ""
                    sam  = str(e.sAMAccountName)    if e.sAMAccountName    else ""
                    name = str(e.displayName)        if e.displayName       else sam
                    mail = str(e.mail)               if e.mail              else ""
                    if dn and sam:
                        results.append((name, sam, dn, mail))
                except Exception:
                    pass
            return sorted(results, key=lambda x: x[0].lower())
        except Exception as ex:
            self._last_error = str(ex)
            return []

    # Country name → ISO 2-letter code map
    COUNTRY_CODES = {
        "india": "IN", "united kingdom": "GB", "uk": "GB",
        "united states": "US", "usa": "US", "us": "US",
        "australia": "AU", "canada": "CA", "germany": "DE",
        "france": "FR", "singapore": "SG", "uae": "AE",
        "japan": "JP", "netherlands": "NL", "sweden": "SE",
        "other": "",
    }

    def get_manager_dn(self, name_or_sam):
        """Resolve a manager by displayName or sAMAccountName to their DN."""
        if not self.connected or not name_or_sam:
            return ""
        # Already a DN?
        if "DC=" in name_or_sam.upper() and "CN=" in name_or_sam.upper():
            return name_or_sam
        try:
            safe = name_or_sam.replace("(", r"\(").replace(")", r"\)")
            self.conn.search(
                search_base=self.base_dn,
                search_filter=f"(|(displayName={safe})(sAMAccountName={safe})(cn={safe}))",
                attributes=["distinguishedName"],
                size_limit=5
            )
            if self.conn.entries:
                return str(self.conn.entries[0].distinguishedName)
        except Exception:
            pass
        return ""

    def create_user(self, ou, fields, group_dns, password=None,
                    account_enabled=True, must_change_pwd=False, pwd_never_expires=False):
        """
        Two-phase creation:
          Phase 1 — create user with ONLY the attributes AD requires at creation time
          Phase 2 — modify to add optional attributes (avoids constraintViolation)
          Phase 3 — set password if provided, then enable account
          Phase 4 — add to groups
        """
        # Build sAMAccountName (max 20 chars, no spaces)
        sam = fields.get("sAMAccountName", "").strip()
        if not sam:
            fn = fields.get("givenName", "")
            ln = fields.get("sn", "")
            sam = (fn[:1] + ln).lower().replace(" ", "")[:20]

        # Build DN — ou is always a full DN from our OUDropdown
        if self.base_dn and self.base_dn in ou:
            dn = f"CN={fields['displayName']},{ou}"
        else:
            dn = f"CN={fields['displayName']},{ou},{self.base_dn}"

        # ── PHASE 1: Minimal required attributes only ─────────────
        # AD constraintViolation is caused by sending optional attributes
        # in the same add() call that conflict with schema/policy.
        # Solution: add ONLY the mandatory attributes first.
        minimal_attrs = {
            "objectClass"       : ["top", "person", "organizationalPerson", "user"],
            "sn"                : fields.get("sn", ""),
            "sAMAccountName"    : sam,
            "userPrincipalName" : fields.get("userPrincipalName", ""),
            "userAccountControl": 514,   # Always start disabled; enable in Phase 3 based on admin choice
        }

        ok = self.conn.add(dn, attributes=minimal_attrs)
        if not ok:
            result = self.conn.result
            code   = result.get("description", "unknown")
            info   = result.get("message", "")
            return False, f"Phase 1 failed [{code}]: {info}"

        errors = []

        # ── PHASE 2: Set optional attributes via modify ───────────
        # Build country code
        country_raw  = fields.get("co", "").strip()
        country_code = self.COUNTRY_CODES.get(country_raw.lower(), country_raw[:2].upper() if country_raw else "")

        optional_map = {
            "displayName"               : fields.get("displayName", ""),
            "givenName"                 : fields.get("givenName", ""),
            "title"                     : fields.get("title", ""),
            "company"                   : fields.get("company", ""),
            "department"                : fields.get("department", ""),
            "physicalDeliveryOfficeName": fields.get("physicalDeliveryOfficeName", ""),
            "streetAddress"             : fields.get("streetAddress", ""),
            "l"                         : fields.get("l", ""),
            "st"                        : fields.get("st", ""),
            "postalCode"                : fields.get("postalCode", ""),
            "telephoneNumber"           : fields.get("telephoneNumber", ""),
            "facsimileTelephoneNumber"  : fields.get("facsimileTelephoneNumber", ""),
        }
        # Add country code only if valid 2-letter code
        if country_code and len(country_code) == 2:
            optional_map["c"]           = country_code
            optional_map["co"]          = country_raw
            optional_map["countryCode"] = str(self._country_to_num(country_code))

        # Mail — add separately as it can conflict in some Exchange schemas
        mail = fields.get("mail", "").strip()

        # Build modify changes dict (only non-empty values)
        changes = {}
        for attr, val in optional_map.items():
            if val and val.strip():
                changes[attr] = [(MODIFY_REPLACE, [val.strip()])]

        if changes:
            mod_ok = self.conn.modify(dn, changes)
            if not mod_ok:
                errors.append(f"Optional attrs warning: {self.conn.result.get('description','')}")

        # Mail — separate modify to isolate Exchange schema issues
        if mail:
            mail_ok = self.conn.modify(dn, {"mail": [(MODIFY_REPLACE, [mail])]})
            if not mail_ok:
                errors.append(f"Mail attr warning: {self.conn.result.get('description','')}")

        # ── PHASE 3: Password, account state, must-change-pwd ───────
        #
        # userAccountControl bit flags:
        #   512   = NORMAL_ACCOUNT (enabled)
        #   514   = NORMAL_ACCOUNT | ACCOUNTDISABLE
        #   66048 = NORMAL_ACCOUNT | DONT_EXPIRE_PASSWORD

        if password:
            try:
                pwd_ok_bool, pwd_method, pwd_err = self.set_password(dn, password)

                if pwd_ok_bool:
                    # Password set — now enable/disable account
                    uac = 512 if account_enabled else 514
                    if pwd_never_expires:
                        uac = uac | 0x10000
                    uac_ok = self.conn.modify(dn, {
                        "userAccountControl": [(MODIFY_REPLACE, [uac])]
                    })
                    if not uac_ok:
                        errors.append(f"UAC set failed: {self.conn.result.get('description','')}")
                    # Force password change at next logon if requested
                    if must_change_pwd and not pwd_never_expires:
                        self.conn.modify(dn, {"pwdLastSet": [(MODIFY_REPLACE, [0])]})
                else:
                    errors.append(
                        f"Password could not be set (all 4 methods failed — "
                        f"unicodePwd, Extended op, dedicated LDAPS port 636, StartTLS). "
                        f"Last error: {pwd_err}. "
                        f"Account is DISABLED. Most likely port 636 is blocked on the DC firewall. "
                        f"Set password via Manage User panel after opening port 636, "
                        f"or set it manually in ADUC."
                    )
            except Exception as pe:
                errors.append(f"Password phase error: {pe}")
        else:
            # No password supplied — AD will keep account disabled regardless of UAC.
            # We still set the UAC value to reflect intent; if admin later sets a
            # password in ADUC, the account will adopt this UAC state.
            uac = 512 if account_enabled else 514
            if pwd_never_expires:
                uac = uac | 0x10000
            self.conn.modify(dn, {"userAccountControl": [(MODIFY_REPLACE, [uac])]})
            errors.append(
                "No password provided — AD keeps the account DISABLED until a password "
                "is set. To enable: open ADUC → right-click user → Reset Password, "
                "then uncheck 'Account is disabled'."
            )

        # ── PHASE 3b: Manager lookup ──────────────────────────────
        manager_raw = fields.get("manager", "").strip()
        if manager_raw:
            mgr_dn = self.get_manager_dn(manager_raw)
            if mgr_dn:
                mgr_ok = self.conn.modify(dn, {"manager": [(MODIFY_REPLACE, [mgr_dn])]})
                if not mgr_ok:
                    errors.append(f"Manager attr failed: {self.conn.result.get('description','')}")
            else:
                errors.append(
                    f"Manager '{manager_raw}' not found in AD — "
                    f"enter the exact Display Name or sAMAccountName of the manager."
                )

        # ── PHASE 4: Add to groups ────────────────────────────────
        grp_errors = []
        for gdn in group_dns:
            try:
                g_ok = self.conn.modify(gdn, {"member": [(MODIFY_ADD, [dn])]})
                if not g_ok:
                    grp_errors.append(gdn.split(",")[0].replace("CN=", ""))
            except Exception:
                grp_errors.append(gdn)
        if grp_errors:
            errors.append(f"Group membership failed for: {', '.join(grp_errors)}")

        warning_str = (" | Warnings: " + "; ".join(errors)) if errors else ""
        _record("create", "user", fields.get("displayName",""))
        return True, f"User '{fields['displayName']}' created successfully.{warning_str} DN: {dn}"

    @staticmethod
    def _country_to_num(iso2):
        """Return numeric country code for common ISO 2-letter codes."""
        codes = {
            "IN":356,"GB":826,"US":840,"AU":36,"CA":124,
            "DE":276,"FR":250,"SG":702,"AE":784,"JP":392,
            "NL":528,"SE":752,
        }
        return codes.get(iso2.upper(), 0)

    # ── GROUPS ───────────────────────────────────────────────────
    def create_group(self, ou, name, group_type, description, member_dns):
        """
        group_type:
          Security   → groupType = -2147483646  (Global Security)
          Distribution → groupType = 2           (Global Distribution)
        """
        # ou from OUDropdown is always a full DN (e.g. OU=Groups,DC=contoso,DC=com).
        # Only append base_dn if ou doesn't already contain it.
        if self.base_dn and self.base_dn.lower() in ou.lower():
            dn = f"CN={name},{ou}"
        else:
            dn = f"CN={name},{ou},{self.base_dn}"
        gtype_map = {
            "Global Security"      : -2147483646,
            "Domain Local Security": -2147483644,
            "Universal Security"   : -2147483640,
            "Global Distribution"  : 2,
            "Domain Local Dist."   : 4,
            "Universal Distribution": 8,
        }
        gtype = gtype_map.get(group_type, -2147483646)
        attrs = {
            "objectClass"   : ["top", "group"],
            "cn"            : name,
            "sAMAccountName": name[:20],
            "groupType"     : gtype,          # integer, not str — ldap3 requirement
        }
        if description:
            attrs["description"] = description
        if member_dns:
            attrs["member"] = member_dns
        ok = self.conn.add(dn, attributes=attrs)
        if not ok:
            result = self.conn.result
            desc   = result.get("description", "unknown")
            info   = result.get("message", "")
            return False, f"Group creation failed [{desc}]: {info} | DN attempted: {dn}"
        _record("create", "group", name)
        return True, f"Group '{name}' created. DN: {dn}"


# ─────────────────────────────────────────────────────────────────────────────
#  REUSABLE WIDGETS
# ─────────────────────────────────────────────────────────────────────────────
class SectionLabel(tk.Frame):
    def __init__(self, parent, text, **kw):
        super().__init__(parent, bg=SURFACE, **kw)
        tk.Label(self, text=text, font=FONT_SUB, fg=ACCENT, bg=SURFACE,
                 anchor="w").pack(fill="x")
        tk.Frame(self, height=1, bg=ACCENT).pack(fill="x", pady=(2, 6))


def labeled_entry(parent, label, required=False, width=35):
    row = tk.Frame(parent, bg=SURFACE)
    lbl = label + ("  *" if required else "")
    tk.Label(row, text=lbl, font=FONT_BODY, fg=TEXT2, bg=SURFACE,
             width=22, anchor="w").pack(side="left")
    var = tk.StringVar()
    ent = tk.Entry(row, textvariable=var, font=FONT_BODY, width=width,
                   relief="solid", bd=1, highlightthickness=1,
                   highlightcolor=ACCENT, bg=SURFACE2)
    ent.pack(side="left", padx=(4, 0))
    row.pack(fill="x", pady=3, padx=8)
    return var, ent


def labeled_combo(parent, label, values, width=33):
    row = tk.Frame(parent, bg=SURFACE)
    tk.Label(row, text=label, font=FONT_BODY, fg=TEXT2, bg=SURFACE,
             width=22, anchor="w").pack(side="left")
    var = tk.StringVar()
    cb  = ttk.Combobox(row, textvariable=var, values=values, width=width,
                       state="readonly", font=FONT_BODY)
    if values:
        cb.current(0)
    cb.pack(side="left", padx=(4, 0))
    row.pack(fill="x", pady=3, padx=8)
    return var, cb


class MultiSelectListbox(tk.Frame):
    """Scrollable multi-select listbox with search filter."""
    def __init__(self, parent, items, label="Select items", height=6, **kw):
        super().__init__(parent, bg=SURFACE, **kw)
        tk.Label(self, text=label, font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE, anchor="w").pack(fill="x", padx=4, pady=(4, 2))
        # search bar
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._filter)
        srch = tk.Entry(self, textvariable=self._search_var, font=FONT_BODY,
                        relief="solid", bd=1, fg=TEXT3,
                        highlightthickness=1, highlightcolor=ACCENT)
        srch.insert(0, "🔍  Search…")
        srch.bind("<FocusIn>",  lambda e: srch.delete(0, "end"))
        srch.pack(fill="x", padx=4, pady=(0, 4))
        # listbox
        frame = tk.Frame(self, bg=SURFACE)
        frame.pack(fill="both", expand=True, padx=4)
        self._lb = tk.Listbox(frame, selectmode="multiple", font=FONT_BODY,
                              height=height, exportselection=False,
                              bd=1, relief="solid", selectbackground=ACCENT,
                              selectforeground=SURFACE, activestyle="none",
                              bg=SURFACE2, fg=TEXT)
        sb = ttk.Scrollbar(frame, orient="vertical", command=self._lb.yview)
        self._lb.configure(yscrollcommand=sb.set)
        self._lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._all_items = items   # list of (display_name, value)
        self._filtered  = list(items)
        self._populate(items)

    def _populate(self, items):
        self._lb.delete(0, "end")
        for display, _ in items:
            self._lb.insert("end", display)

    def _filter(self, *_):
        q = self._search_var.get().strip().lower()
        if q and q != "🔍  search…" and q != "🔍  search...":
            filtered = [(d, v) for d, v in self._all_items if q in d.lower()]
        else:
            filtered = self._all_items
        self._populate(filtered)
        self._filtered = filtered

    def update_items(self, items):
        self._all_items = items
        self._filtered  = list(items)   # always sync _filtered so selections work immediately
        self._populate(items)

    def get_selected_values(self):
        """Return list of values for selected indices."""
        sel = self._lb.curselection()
        source = getattr(self, "_filtered", self._all_items)
        return [source[i][1] for i in sel if i < len(source)]

    def get_selected_labels(self):
        sel = self._lb.curselection()
        source = getattr(self, "_filtered", self._all_items)
        return [source[i][0] for i in sel if i < len(source)]



class OUDropdown(tk.Frame):
    """Searchable OU selector with popup listbox."""
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=SURFACE, **kw)
        # Top row: label + refresh button
        top = tk.Frame(self, bg=SURFACE)
        top.pack(fill="x", padx=8, pady=(4, 2))
        tk.Label(top, text="Target OU  *", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE, anchor="w", width=22).pack(side="left")
        self._refresh_btn = tk.Button(top, text="↻ Refresh OUs", font=FONT_TINY,
                                      bg=LIGHT_BLU, fg=ACCENT, relief="flat",
                                      padx=8, pady=2, cursor="hand2")
        self._refresh_btn.pack(side="right")

        # Search entry
        row = tk.Frame(self, bg=SURFACE)
        row.pack(fill="x", padx=8, pady=(0, 2))
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._on_type)
        self._entry = tk.Entry(row, textvariable=self._search_var,
                               font=FONT_BODY, width=50, relief="solid", bd=1,
                               highlightthickness=1, highlightcolor=ACCENT,
                               bg=SURFACE2, fg=TEXT)
        self._entry.pack(fill="x", expand=True)
        self._entry.bind("<FocusIn>",  self._on_focus_in)
        self._entry.bind("<Return>",   lambda e: self._hide_popup())

        # Inline listbox (always visible when OUs are loaded)
        lf = tk.Frame(self, bg=SURFACE)
        lf.pack(fill="x", padx=8, pady=(0, 4))
        self._lb = tk.Listbox(lf, font=FONT_BODY, height=5,
                              bd=1, relief="solid", exportselection=False,
                              selectbackground=ACCENT, selectforeground=SURFACE,
                              activestyle="none", bg=SURFACE2, fg=TEXT)
        sb = ttk.Scrollbar(lf, orient="vertical", command=self._lb.yview)
        self._lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._lb.pack(side="left", fill="both", expand=True)
        self._lb.bind("<<ListboxSelect>>", self._on_select)

        # Selection indicator
        self._info = tk.Label(self, font=FONT_TINY, fg=TEAL, bg=LIGHT_BLU,
                              anchor="w", padx=8, pady=3,
                              text="No OU selected — click Refresh OUs, then pick one")
        self._info.pack(fill="x", padx=8, pady=(0, 4))

        self._all_ous  = []
        self._filtered = []
        self._sel_dn   = ""
        self._popup    = None

    def _on_focus_in(self, e):
        pass  # inline listbox always visible

    def _on_type(self, *_):
        q = self._search_var.get().strip().lower()
        self._filtered = [(l, d) for l, d in self._all_ous if q in l.lower()] if q else list(self._all_ous)
        self._lb.delete(0, "end")
        for label, _ in self._filtered:
            self._lb.insert("end", label)

    def _on_select(self, e):
        sel = self._lb.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx < len(self._filtered):
            label, dn = self._filtered[idx]
            self._sel_dn = dn
            self._search_var.set(label)
            self._info.config(
                text=f"Selected OU:  {dn}",
                fg=TEAL, bg=LIGHT_BLU)

    def _hide_popup(self):
        pass

    def update_ous(self, ou_list):
        self._all_ous  = ou_list
        self._filtered = list(ou_list)
        self._lb.delete(0, "end")
        for label, _ in ou_list:
            self._lb.insert("end", label)
        count = len(ou_list)
        self._info.config(
            text=f"{count} OU(s) loaded — search above or click to select",
            fg=ACCENT, bg=LIGHT_BLU)

    def get_ou_dn(self):
        return self._sel_dn.strip()

    def set_refresh_command(self, cmd):
        self._refresh_btn.config(command=cmd)

    def reset(self):
        self._sel_dn = ""
        self._search_var.set("")
        self._lb.selection_clear(0, "end")
        self._info.config(
            text="No OU selected — click Refresh OUs, then pick one",
            fg=TEXT2, bg=LIGHT_BLU)


class StatusBar(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=NAVY, height=28, **kw)
        self._var = tk.StringVar(value="Ready")
        tk.Label(self, textvariable=self._var, font=FONT_TINY,
                 fg="white", bg=NAVY, padx=10).pack(side="left")
        self._dot = tk.Label(self, text="●", font=FONT_TINY,
                             fg="#94A3B8", bg=NAVY, padx=6)
        self._dot.pack(side="right")

    def set(self, msg, kind="info"):
        self._var.set(msg)
        colours = {"ok": "#10d98a", "error": "#FF5F6E", "info": "#93C5FD", "warn": "#FBBF24"}
        self._dot.config(fg=colours.get(kind, "#93C5FD"))


class ActivityLog:
    """
    Lightweight logger — not a visible widget itself.
    Messages are written to any ScrolledText registered via register_widget().
    This keeps the Activity Log out of every panel and only in the Log tab.
    """
    def __init__(self):
        self._widgets = []   # list of ScrolledText widgets to write to
        self._buffer  = []   # store messages so new widgets catch up

    def register_widget(self, txt_widget):
        """Register a ScrolledText; replay buffered messages into it."""
        txt_widget.tag_config("ok",    foreground="#10d98a")
        txt_widget.tag_config("error", foreground="#FF5F6E")
        txt_widget.tag_config("info",  foreground="#93C5FD")
        txt_widget.tag_config("warn",  foreground="#FBBF24")
        txt_widget.tag_config("ts",    foreground="#475569")
        # Replay history
        txt_widget.config(state="normal")
        for ts, msg, kind in self._buffer:
            txt_widget.insert("end", f"[{ts}] ", "ts")
            txt_widget.insert("end", msg + "\n", kind)
        txt_widget.see("end")
        txt_widget.config(state="disabled")
        self._widgets.append(txt_widget)

    def log(self, msg, kind="info"):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self._buffer.append((ts, msg, kind))
        if len(self._buffer) > 500:
            self._buffer = self._buffer[-500:]
        for txt in self._widgets:
            try:
                txt.config(state="normal")
                txt.insert("end", f"[{ts}] ", "ts")
                txt.insert("end", msg + "\n", kind)
                txt.see("end")
                txt.config(state="disabled")
            except Exception:
                pass


def _validate_base_dn(dn: str):
    """
    Returns (is_valid: bool, suggestion: str).
    A valid Base DN looks like:  DC=domain,DC=tld
    Common mistakes caught:
      • DC=,DC=xxx      — first DC component is blank
      • domain.com      — dot-notation instead of DN notation
      • empty string
    """
    dn = dn.strip()
    if not dn:
        return False, "Base DN cannot be empty. Example: DC=contoso,DC=com"
    # Dot-notation shortcut — auto-convert domain.com → DC=domain,DC=com
    if re.match(r'^[a-zA-Z0-9\-]+(\.[a-zA-Z0-9\-]+)+$', dn):
        converted = ",".join(f"DC={part}" for part in dn.split("."))
        return False, f"Use DN format, not dot notation. Did you mean: {converted}"
    # Must contain at least one DC= component
    if "DC=" not in dn.upper():
        return False, "Base DN must contain DC= components. Example: DC=contoso,DC=com"
    # Each DC= component must have a non-empty value
    for part in dn.split(","):
        part = part.strip()
        if part.upper().startswith("DC=") and len(part) <= 3:
            return False, f"Empty DC component found in '{dn}'. Example: DC=contoso,DC=com"
    return True, ""


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN / CONNECTION WINDOW
# ─────────────────────────────────────────────────────────────────────────────
class LoginWindow(tk.Toplevel):
    def __init__(self, master, ad: ADConnector, on_success):
        super().__init__(master)
        self.title("Active Directory — Authenticate")
        self.resizable(False, False)
        self.configure(bg=SURFACE)
        self.grab_set()
        self.ad = ad
        self.on_success = on_success
        self._build()
        self.after(100, self._center)

    def _center(self):
        self.update_idletasks()
        w = max(self.winfo_width(), 600)
        h = max(self.winfo_height(), 420)
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        # ── Load saved config ─────────────────────────────────────
        cfg = _load_config()
        act_ok, act_rem, act_exp = _get_activation_status(cfg)

        saved_host    = cfg.get("ad_host",  "dc.company.com")
        saved_port    = cfg.get("ad_port",  "389")
        saved_dn      = cfg.get("base_dn",  "DC=company,DC=com")
        saved_user    = cfg.get("username", "admin@company.com")
        saved_ssl_raw = cfg.get("use_ssl", False)
        saved_ssl     = saved_ssl_raw and str(saved_port) == "636"

        # ── Colour palette (login-specific extras) ────────────────
        HDR_TOP    = "#0A1C40"   # deep navy top of header
        HDR_BOT    = "#1245A8"   # vivid blue bottom of header
        CARD_BG    = "#F7F9FC"   # light card background
        BADGE_OK   = "#065F46"   # activated badge bg
        BADGE_WARN = "#78350F"   # not-activated badge bg
        DIVIDER    = "#DDE3EE"
        FIELD_BG   = "#FFFFFF"
        FIELD_BD   = "#C7D2E2"
        FIELD_ACT  = "#1D6FD8"

        self.configure(bg=CARD_BG)

        # ══════════════════════════════════════════════════════════
        # HEADER  — two-tone banner with logo area + tagline
        # ══════════════════════════════════════════════════════════
        hdr = tk.Frame(self, bg=HDR_TOP, pady=0)
        hdr.pack(fill="x")

        # Top strip — icon + title + activation badge
        top_strip = tk.Frame(hdr, bg=HDR_TOP, padx=16, pady=8)
        top_strip.pack(fill="x")

        icon_lbl = tk.Label(top_strip, text="🏢", font=("Arial", 18),
                            bg=HDR_TOP, fg=WHITE)
        icon_lbl.pack(side="left")

        title_block = tk.Frame(top_strip, bg=HDR_TOP)
        title_block.pack(side="left", padx=(8, 0))
        tk.Label(title_block, text="Identity Manager",
                 font=("Arial", 13, "bold"), fg=WHITE, bg=HDR_TOP,
                 anchor="w").pack(anchor="w")
        tk.Label(title_block, text="Active Directory Administration",
                 font=("Arial", 8), fg="#93C5FD", bg=HDR_TOP,
                 anchor="w").pack(anchor="w")

        # Activation badge
        act_badge_bg  = BADGE_OK   if act_ok else BADGE_WARN
        act_badge_txt = (f"✓  {act_rem}d left" if act_ok else "⚠  Not activated")
        act_badge_fg  = "#6EE7B7" if act_ok else "#FDE68A"
        badge_frame = tk.Frame(top_strip, bg=act_badge_bg, padx=8, pady=3)
        badge_frame.pack(side="right", anchor="n", pady=2)
        tk.Label(badge_frame, text=act_badge_txt,
                 font=("Arial", 7, "bold"), fg=act_badge_fg,
                 bg=act_badge_bg).pack()
        if act_ok:
            tk.Label(badge_frame, text=f"Expires {act_exp}",
                     font=("Arial", 6), fg="#A7F3D0", bg=act_badge_bg).pack()

        # Accent stripe below the banner
        tk.Frame(hdr, bg=HDR_BOT, height=4).pack(fill="x")

        # ── Vendor sub-bar ────────────────────────────────────────
        vendor_bar = tk.Frame(self, bg="#E8EEF8", padx=16, pady=3)
        vendor_bar.pack(fill="x")
        tk.Label(vendor_bar, text=f"Powered by  {_VENDOR_NAME}",
                 font=("Arial", 7), fg="#4B6089", bg="#E8EEF8").pack(side="left")
        tk.Label(vendor_bar, text=f"v{_APP_VERSION}",
                 font=("Arial", 7, "bold"), fg="#4B6089", bg="#E8EEF8").pack(side="right")

        # ══════════════════════════════════════════════════════════
        # BODY  — two-column card layout
        # ══════════════════════════════════════════════════════════
        body_outer = tk.Frame(self, bg=CARD_BG, padx=12, pady=8)
        body_outer.pack(fill="both", expand=True)

        # Left column: connection fields card
        left_card = tk.Frame(body_outer, bg=WHITE, relief="flat",
                             highlightthickness=1,
                             highlightbackground=DIVIDER)
        left_card.pack(side="left", fill="both", expand=True,
                       padx=(0, 8), pady=0)

        # Right column: SSL + notes card
        right_card = tk.Frame(body_outer, bg=WHITE, relief="flat",
                              highlightthickness=1,
                              highlightbackground=DIVIDER)
        right_card.pack(side="left", fill="both", expand=False, pady=0)
        right_card.configure(width=190)
        right_card.pack_propagate(False)

        # ── LEFT CARD: section header ─────────────────────────────
        lc_hdr = tk.Frame(left_card, bg=NAVY, padx=10, pady=5)
        lc_hdr.pack(fill="x")
        tk.Label(lc_hdr, text="🔌  Server Connection",
                 font=("Arial", 9, "bold"), fg=WHITE, bg=NAVY,
                 anchor="w").pack(anchor="w")

        lc_body = tk.Frame(left_card, bg=WHITE, padx=12, pady=6)
        lc_body.pack(fill="both", expand=True)

        # Helper: styled entry row
        self._vars    = {}
        self._entries = {}

        def _field(parent, label, key, default, secret=False, hint=""):
            row = tk.Frame(parent, bg=WHITE)
            row.pack(fill="x", pady=(0, 6))
            tk.Label(row, text=label, font=("Arial", 7, "bold"),
                     fg="#5A6A85", bg=WHITE, anchor="w").pack(anchor="w")
            var = tk.StringVar(value=default)
            ent = tk.Entry(row, textvariable=var,
                           font=("Arial", 9), width=30,
                           show="*" if secret else "",
                           relief="flat", bd=0,
                           bg=FIELD_BG, fg=TEXT,
                           highlightthickness=1,
                           highlightbackground=FIELD_BD,
                           highlightcolor=FIELD_ACT,
                           insertbackground=ACCENT)
            ent.pack(fill="x", ipady=4)
            if hint:
                tk.Label(row, text=hint, font=("Arial", 6),
                         fg=TEXT3, bg=WHITE, anchor="w").pack(anchor="w")
            self._vars[key]    = var
            self._entries[key] = ent
            return var, ent

        _field(lc_body, "AD SERVER / DOMAIN CONTROLLER", "ad_host", saved_host,
               hint="e.g. dc01.contoso.com")
        _field(lc_body, "LDAP PORT", "ad_port", saved_port,
               hint="389 = LDAP (default)  ·  636 = LDAPS")
        _field(lc_body, "BASE DN", "base_dn", saved_dn,
               hint="e.g. DC=contoso,DC=com")
        _field(lc_body, "USERNAME  (UPN or DOMAIN\\user)", "username", saved_user,
               hint="e.g. admin@contoso.com")
        _field(lc_body, "PASSWORD", "password", "", secret=True)

        # Live Base DN validation hint (inline below the field)
        self._dn_hint = tk.Label(lc_body, text="", font=("Arial", 7),
                                 fg="#F59E0B", bg=WHITE, anchor="w")
        self._dn_hint.pack(fill="x", pady=(0, 4))

        def _check_dn(*_):
            dn = self._vars["base_dn"].get().strip()
            ok, hint = _validate_base_dn(dn)
            if not ok and dn:
                self._dn_hint.config(text=f"⚠  {hint}", fg=RED)
                self._entries["base_dn"].config(highlightbackground=RED,
                                                highlightcolor=RED,
                                                highlightthickness=2)
            else:
                self._dn_hint.config(
                    text="✓  Base DN format looks correct." if dn else "",
                    fg="#059669")
                self._entries["base_dn"].config(highlightbackground=FIELD_BD,
                                                highlightcolor=FIELD_ACT,
                                                highlightthickness=1)
        self._vars["base_dn"].trace_add("write", _check_dn)
        _check_dn()

        # ── RIGHT CARD ────────────────────────────────────────────
        rc_hdr = tk.Frame(right_card, bg="#1245A8", padx=10, pady=5)
        rc_hdr.pack(fill="x")
        tk.Label(rc_hdr, text="⚙  Options",
                 font=("Arial", 9, "bold"), fg=WHITE, bg="#1245A8",
                 anchor="w").pack(anchor="w")

        rc_body = tk.Frame(right_card, bg=WHITE, padx=10, pady=8)
        rc_body.pack(fill="both", expand=True)

        # SSL toggle
        self._ssl = tk.BooleanVar(value=saved_ssl)

        ssl_card = tk.Frame(rc_body, bg="#EEF4FF", padx=8, pady=6,
                            highlightthickness=1,
                            highlightbackground="#BFD0F0")
        ssl_card.pack(fill="x", pady=(0, 7))
        ssl_top = tk.Frame(ssl_card, bg="#EEF4FF")
        ssl_top.pack(fill="x")
        tk.Label(ssl_top, text="🔒  Use SSL (LDAPS)",
                 font=("Arial", 8, "bold"), fg=NAVY, bg="#EEF4FF").pack(side="left")

        def _on_ssl_toggle():
            if self._ssl.get():
                self._vars["ad_port"].set("636")
            else:
                if self._vars["ad_port"].get().strip() == "636":
                    self._vars["ad_port"].set("389")
        tk.Checkbutton(ssl_top, variable=self._ssl, bg="#EEF4FF",
                       activebackground="#EEF4FF",
                       command=_on_ssl_toggle).pack(side="right")
        tk.Label(ssl_card,
                 text="Port 636 · full encryption.\nRecommended if DC supports it.",
                 font=("Arial", 6), fg="#4B6089", bg="#EEF4FF",
                 justify="left").pack(anchor="w", pady=(3, 0))

        # Info note
        info_card = tk.Frame(rc_body, bg="#F0FDF4", padx=8, pady=6,
                             highlightthickness=1,
                             highlightbackground="#BBF7D0")
        info_card.pack(fill="x", pady=(0, 7))
        tk.Label(info_card, text="ℹ  Plain LDAP (389)",
                 font=("Arial", 7, "bold"), fg="#065F46", bg="#F0FDF4").pack(anchor="w")
        tk.Label(info_card,
                 text="Passwords via StartTLS\nautomatically on most DCs.",
                 font=("Arial", 6), fg="#15803D", bg="#F0FDF4",
                 justify="left").pack(anchor="w", pady=(2, 0))

        # Demo note
        demo_card = tk.Frame(rc_body, bg="#FFFBEB", padx=8, pady=6,
                             highlightthickness=1,
                             highlightbackground="#FDE68A")
        demo_card.pack(fill="x")
        tk.Label(demo_card, text="💡  Demo Mode",
                 font=("Arial", 7, "bold"), fg="#92400E", bg="#FFFBEB").pack(anchor="w")
        tk.Label(demo_card,
                 text="No AD needed.\nExplore with simulated data.",
                 font=("Arial", 6), fg="#B45309", bg="#FFFBEB",
                 justify="left").pack(anchor="w", pady=(2, 0))

        # ══════════════════════════════════════════════════════════
        # FOOTER  — action buttons + status
        # ══════════════════════════════════════════════════════════
        footer = tk.Frame(self, bg="#E8EEF8", padx=14, pady=8)
        footer.pack(fill="x")

        btn_connect = tk.Button(
            footer, text="  🔗  Connect to AD  ",
            font=("Arial", 9, "bold"), bg=ACCENT, fg=WHITE,
            relief="flat", padx=12, pady=6, cursor="hand2",
            activebackground=ACCENT2, activeforeground=WHITE,
            command=self._connect)
        btn_connect.pack(side="left", padx=(0, 6))
        btn_connect.bind("<Enter>", lambda e: btn_connect.config(bg=ACCENT2))
        btn_connect.bind("<Leave>", lambda e: btn_connect.config(bg=ACCENT))

        btn_ldaps = tk.Button(
            footer, text="🔒 Test LDAPS",
            font=("Arial", 8), bg=AMBER, fg=WHITE,
            relief="flat", padx=8, pady=6, cursor="hand2",
            activebackground="#B45309", activeforeground=WHITE,
            command=self._test_ldaps)
        btn_ldaps.pack(side="left", padx=(0, 5))
        btn_ldaps.bind("<Enter>", lambda e: btn_ldaps.config(bg="#B45309"))
        btn_ldaps.bind("<Leave>", lambda e: btn_ldaps.config(bg=AMBER))

        btn_demo = tk.Button(
            footer, text="🧪 Demo Mode",
            font=("Arial", 8), bg=TEAL, fg=WHITE,
            relief="flat", padx=8, pady=6, cursor="hand2",
            activebackground="#0A5C50", activeforeground=WHITE,
            command=self._demo)
        btn_demo.pack(side="left", padx=(0, 5))
        btn_demo.bind("<Enter>", lambda e: btn_demo.config(bg="#0A5C50"))
        btn_demo.bind("<Leave>", lambda e: btn_demo.config(bg=TEAL))

        btn_cancel = tk.Button(
            footer, text="✕ Cancel",
            font=("Arial", 8), bg="#94A3B8", fg=WHITE,
            relief="flat", padx=8, pady=6, cursor="hand2",
            activebackground="#64748B", activeforeground=WHITE,
            command=self.destroy)
        btn_cancel.pack(side="left")
        btn_cancel.bind("<Enter>", lambda e: btn_cancel.config(bg="#64748B"))
        btn_cancel.bind("<Leave>", lambda e: btn_cancel.config(bg="#94A3B8"))

        self._status_lbl = tk.Label(
            self, text="", font=("Arial", 7),
            fg=RED, bg=CARD_BG, wraplength=560,
            justify="left", padx=14, pady=4)
        self._status_lbl.pack(fill="x")

    def _connect(self):
        host = self._vars["ad_host"].get().strip()
        port = self._vars["ad_port"].get().strip() or "389"
        dn   = self._vars["base_dn"].get().strip()
        user = self._vars["username"].get().strip()
        pwd  = self._vars["password"].get()
        ssl  = self._ssl.get()

        # Validate Base DN before attempting connection
        dn_ok, dn_hint = _validate_base_dn(dn)
        if not dn_ok:
            self._status_lbl.config(
                text=f"✗ Invalid Base DN — {dn_hint}", fg=RED)
            self._entries["base_dn"].focus()
            return

        self._status_lbl.config(text="Connecting…", fg=ACCENT)
        self.update()
        ok, msg = self.ad.connect(host=host, port=port, base_dn=dn,
                                  username=user, password=pwd, use_ssl=ssl)
        if ok:
            # Save config — use the actual ssl mode that succeeded, not what was requested
            cfg = _load_config()
            cfg.update({"ad_host": host, "ad_port": str(self.ad._port),
                        "base_dn": dn, "use_ssl": self.ad.use_ssl, "username": user})
            _save_config(cfg)
            # Reflect any auto-fallback in the UI fields
            self._ssl.set(self.ad.use_ssl)
            self._vars["ad_port"].set(str(self.ad._port))
            self._status_lbl.config(text="✓ " + msg, fg=GREEN)
            self.after(600, lambda: [self.destroy(), self.on_success(real=True)])
        else:
            # Reset SSL checkbox to unchecked on any SSL failure so the user
            # doesn't keep retrying with a bad SSL setting
            if ssl and ("ssl" in msg.lower() or "tls" in msg.lower()
                        or "10054" in msg or "startTLS" in msg.lower()):
                self._ssl.set(False)
                if self._vars["ad_port"].get().strip() == "636":
                    self._vars["ad_port"].set("389")
            self._status_lbl.config(text=f"✗  {msg}",
                                    fg=RED, wraplength=460, justify="left")

    def _test_ldaps(self):
        """Quick test: can we open a TCP socket to port 636 on the DC?"""
        host = self._vars["ad_host"].get().strip()
        if not host:
            self._status_lbl.config(text="Enter the AD Server hostname first.", fg=AMBER)
            return
        self._status_lbl.config(text=f"Testing LDAPS (port 636) on {host}…", fg=ACCENT)
        self.update()
        import socket
        try:
            sock = socket.create_connection((host, 636), timeout=5)
            sock.close()
            # Port reachable — now try an actual SSL handshake
            try:
                import ssl as _ssl2
                ctx = _ssl2.SSLContext(_ssl2.PROTOCOL_TLS_CLIENT)
                ctx.check_hostname = False
                ctx.verify_mode    = _ssl2.CERT_NONE
                with ctx.wrap_socket(socket.create_connection((host, 636), timeout=5),
                                     server_hostname=host) as ssock:
                    cert = ssock.getpeercert(binary_form=True)
                self._status_lbl.config(
                    text=f"✓ LDAPS port 636 is OPEN and SSL handshake succeeded on {host}.\n"
                         f"  You can connect with SSL checked + port 636 for full LDAPS support.",
                    fg=GREEN)
            except Exception as ssl_err:
                self._status_lbl.config(
                    text=f"⚠ Port 636 is reachable on {host} but SSL handshake failed: {ssl_err}\n"
                         f"  The DC may not have an LDAPS certificate installed.",
                    fg=AMBER)
        except (socket.timeout, ConnectionRefusedError, OSError) as e:
            self._status_lbl.config(
                text=f"✗ Port 636 is BLOCKED or unreachable on {host}: {e}\n"
                     f"  Use plain LDAP (port 389, SSL unchecked) — passwords will use StartTLS.\n"
                     f"  To enable LDAPS: install a certificate on the DC or open port 636 in the firewall.",
                fg=RED)

    def _demo(self):
        self.ad.connected = True
        self.ad.base_dn   = self._vars["base_dn"].get().strip() or "DC=demo,DC=local"
        self.destroy()
        self.on_success(real=False)


# ─────────────────────────────────────────────────────────────────────────────
#  USER CREATION FORM
# ─────────────────────────────────────────────────────────────────────────────
class UserCreationForm(tk.Frame):
    def __init__(self, parent, ad: ADConnector, logger: ActivityLog, status: StatusBar):
        super().__init__(parent, bg=SURFACE)
        self.ad      = ad
        self.log     = logger
        self.status  = status
        self._groups = []   # (label, dn)
        self._ous    = []   # (label, dn)
        self._build()

    def refresh_ous(self):
        if self.ad.connected:
            self._ous = self.ad.get_all_ous()
            self._ou_dropdown.update_ous(self._ous)
            count = len(self._ous)
            if count <= 1:  # only root or empty
                err = getattr(self.ad, "_last_error", "")
                err_hint = f" Error: {err}" if err else " Check AD permissions — account needs 'Read' on OU objects."
                self.log.log(f"OUs loaded: {count} OU(s). Only root DN available.{err_hint}", "warn")
                self.status.set(f"Only root OU found. {err_hint}", "warn")
            else:
                self.log.log(f"OUs loaded: {count} OU(s) found in AD.", "ok")
                self.status.set(f"{count} OU(s) available — select target OU for new user.", "ok")

    def refresh_users_for_manager(self):
        """Fetch all users and populate the manager searchable dropdown."""
        if self.ad.connected:
            users = self.ad.get_all_users()
            self._mgr_populate_users(users)
            if users:
                self.log.log(f"Manager list loaded: {len(users)} user(s).", "ok")
            else:
                err = getattr(self.ad, "_last_error", "")
                err_hint = f" AD error: {err}" if err else " Check account has Read access to user objects."
                self.log.log(f"Manager list: 0 users found.{err_hint}", "warn")

    def refresh_groups(self):
        if self.ad.connected:
            self._groups = self.ad.get_all_groups()
            self._grp_selector.update_items(self._groups)
            count = len(self._groups)
            if count == 0:
                err = getattr(self.ad, "_last_error", "")
                err_hint = f" AD error: {err}" if err else " Ensure the connecting account has Read access to group objects."
                self.log.log(f"No groups found.{err_hint}", "warn")
                self.status.set(f"No groups found.{err_hint}", "warn")
            else:
                self.log.log(f"Groups loaded: {count} group(s) found in AD.", "ok")
                self.status.set(f"{count} group(s) loaded into Group Membership list.", "ok")

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=LIGHT_BLU, padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="👤  Create User Account",
                 font=FONT_HEAD, fg=NAVY, bg=LIGHT_BLU).pack(side="left")
        tk.Label(hdr, text="Fields marked * are required",
                 font=FONT_TINY, fg=TEXT2, bg=LIGHT_BLU).pack(side="right")

        # Scrollable body
        canvas = tk.Canvas(self, bg=SURFACE, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._inner = tk.Frame(canvas, bg=SURFACE)
        self._win   = canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._win, width=e.width))
        _bind_mousewheel(canvas)

        f = self._inner
        pad = {"padx": 16, "pady": 4, "fill": "x"}

        # ── OU Dropdown ───────────────────────────────────────────
        SectionLabel(f, "Organisational Unit (OU)").pack(**pad)
        self._ou_dropdown = OUDropdown(f)
        self._ou_dropdown.pack(fill="x", padx=8, pady=4)
        self._ou_dropdown.set_refresh_command(self.refresh_ous)

        # ── Identity ─────────────────────────────────────────────
        SectionLabel(f, "Identity").pack(**pad)
        self._display_var,  _ = labeled_entry(f, "Display Name",   required=True)
        self._first_var,    _ = labeled_entry(f, "First Name",     required=True)
        self._last_var,     _ = labeled_entry(f, "Last Name",      required=True)
        self._upn_var,      _ = labeled_entry(f, "User Principal Name (UPN)", required=True, width=40)
        self._sam_var,      _ = labeled_entry(f, "Logon Name (sAMAccountName)", width=25)

        # Password row — show/hide toggle + auto-generate
        pwd_outer = tk.Frame(f, bg=SURFACE2, padx=12, pady=10, relief="solid", bd=1)
        pwd_outer.pack(fill="x", padx=16, pady=(0, 4))

        pwd_row = tk.Frame(pwd_outer, bg=SURFACE2)
        pwd_row.pack(fill="x")
        tk.Label(pwd_row, text="Initial Password", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE2, width=22, anchor="w").pack(side="left")
        self._pwd_var  = tk.StringVar()
        self._pwd_show = tk.BooleanVar(value=False)
        self._pwd_ent  = tk.Entry(pwd_row, textvariable=self._pwd_var, show="*",
                                  font=FONT_BODY, width=26, relief="solid", bd=1,
                                  highlightthickness=1, highlightcolor=ACCENT, bg=SURFACE)
        self._pwd_ent.pack(side="left", padx=(4, 6))

        def _toggle_show():
            self._pwd_ent.config(show="" if self._pwd_show.get() else "*")
        tk.Checkbutton(pwd_row, text="Show", variable=self._pwd_show,
                       font=FONT_TINY, bg=SURFACE2, fg=TEXT2,
                       activebackground=SURFACE2, cursor="hand2",
                       command=_toggle_show).pack(side="left", padx=(0, 8))

        import string, secrets as _sec
        def _gen_password():
            """Generate a strong 12-char password AD will accept."""
            alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
            while True:
                pwd = "".join(_sec.choice(alphabet) for _ in range(12))
                # AD complexity: upper + lower + digit + special
                if (any(c.isupper() for c in pwd) and any(c.islower() for c in pwd)
                        and any(c.isdigit() for c in pwd)
                        and any(c in "!@#$%^&*" for c in pwd)):
                    break
            self._pwd_var.set(pwd)
            self._pwd_show.set(True)
            self._pwd_ent.config(show="")
            _update_pwd_warning()

        tk.Button(pwd_row, text="⚙ Auto-generate", font=FONT_TINY,
                  bg=ACCENT, fg=WHITE, relief="flat", padx=8, pady=3,
                  cursor="hand2", command=_gen_password).pack(side="left")

        # Dynamic password warning — updates as user types
        self._pwd_warn = tk.Label(pwd_outer,
            text="", font=FONT_TINY, fg=AMBER, bg=SURFACE2,
            wraplength=560, justify="left", anchor="w")
        self._pwd_warn.pack(fill="x", pady=(6, 0))

        def _update_pwd_warning(*_):
            pwd = self._pwd_var.get().strip()
            is_ssl = getattr(self.ad, "use_ssl", False)
            if not pwd:
                self._pwd_warn.config(
                    fg=RED,
                    text="⚠  No password set — AD will create the account as DISABLED regardless "
                         "of the 'Account is enabled' checkbox below. You MUST set a password to "
                         "create an enabled account. Use 'Auto-generate' or type one above.")
            elif is_ssl:
                self._pwd_warn.config(
                    fg="#166534",
                    text="✓  Password set. LDAPS (SSL) connection active — password will be "
                         "applied and account will be enabled.")
            else:
                self._pwd_warn.config(
                    fg="#92400E",
                    text="⚠  Password set, but connection is plain LDAP (port 389, no SSL). "
                         "The app will try the Password Modify Extended Operation first (works "
                         "over plain LDAP on most AD environments). If that fails, reconnect "
                         "using LDAPS (port 636 + SSL checkbox) to guarantee password is set.")
        self._pwd_var.trace_add("write", _update_pwd_warning)
        _update_pwd_warning()  # run once on load
        # expose for external call after connection state changes
        self._update_pwd_warning = _update_pwd_warning

        # auto-fill display name
        def _auto_display(*_):
            fn = self._first_var.get().strip()
            ln = self._last_var.get().strip()
            if fn or ln:
                self._display_var.set(f"{fn} {ln}".strip())
            # auto UPN
            upn = self._upn_var.get()
            if "@" not in upn and fn:
                self._sam_var.set((fn[0]+ln).lower()[:20])
        self._first_var.trace_add("write", _auto_display)
        self._last_var.trace_add("write",  _auto_display)

        # ── Job Info ─────────────────────────────────────────────
        SectionLabel(f, "Job Information").pack(**pad)
        self._title_var,   _ = labeled_entry(f, "Job Title")
        self._company_var, _ = labeled_entry(f, "Company Name")
        self._dept_var,    _ = labeled_entry(f, "Department")
        self._office_var,  _ = labeled_entry(f, "Office Location")

        # Manager — searchable dropdown from AD users
        mgr_outer = tk.Frame(f, bg=SURFACE)
        mgr_outer.pack(fill="x", pady=3, padx=8)
        tk.Label(mgr_outer, text="Manager", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE, width=22, anchor="w").pack(side="left")
        mgr_inner = tk.Frame(mgr_outer, bg=SURFACE)
        mgr_inner.pack(side="left", fill="x", expand=True)

        self._mgr_search_var = tk.StringVar()
        self._mgr_dn         = ""        # resolved DN of selected manager
        self._mgr_all_users  = []        # [(display_name, dn), …]
        self._mgr_filtered   = []

        self._mgr_entry = tk.Entry(mgr_inner, textvariable=self._mgr_search_var,
                                   font=FONT_BODY, width=34, relief="solid", bd=1,
                                   highlightthickness=1, highlightcolor=ACCENT, bg=SURFACE2)
        self._mgr_entry.pack(side="left")
        self._mgr_entry.bind("<KeyRelease>", self._mgr_filter)
        self._mgr_entry.bind("<FocusIn>",    self._mgr_show_popup)
        self._mgr_entry.bind("<FocusOut>",   lambda e: self.after(200, self._mgr_hide_popup))
        self._mgr_entry.insert(0, "Type to search manager…")
        self._mgr_entry.config(fg=TEXT3)
        def _mgr_focus_in(e):
            cur = self._mgr_search_var.get()
            if cur == "Type to search manager…" or cur == "":
                self._mgr_entry.delete(0, "end")
                self._mgr_entry.config(fg=TEXT)
            self._mgr_show_popup(e)
        self._mgr_entry.bind("<FocusIn>", _mgr_focus_in, add=True)

        tk.Button(mgr_inner, text="✕", font=FONT_TINY, bg=SURFACE, fg=TEXT3,
                  relief="flat", padx=4, cursor="hand2",
                  command=self._mgr_clear).pack(side="left", padx=(4,0))

        self._mgr_info = tk.Label(mgr_inner, text="No manager selected",
                                  font=FONT_TINY, fg=TEXT3, bg=SURFACE)
        self._mgr_info.pack(side="left", padx=(8, 0))

        self._mgr_popup_win  = None
        self._mgr_popup_lb   = None
        self._mgr_var        = self._mgr_search_var  # alias used in _submit

        # Also keep backward-compat _mgr_var for _reset
        

        # ── Address ──────────────────────────────────────────────
        SectionLabel(f, "Address").pack(**pad)
        self._street_var,  _ = labeled_entry(f, "Street Address", width=40)
        self._city_var,    _ = labeled_entry(f, "City")
        self._state_var,   _ = labeled_entry(f, "State or Province")
        self._zip_var,     _ = labeled_entry(f, "Zip / Postal Code")

        country_list = [
            "", "India", "United Kingdom", "United States", "Australia",
            "Canada", "Germany", "France", "Singapore", "UAE", "Japan",
            "Netherlands", "Sweden", "Other"
        ]
        self._country_var, _ = labeled_combo(f, "Country or Region", country_list)

        # ── Contact ──────────────────────────────────────────────
        SectionLabel(f, "Contact").pack(**pad)
        self._phone_var, _ = labeled_entry(f, "Business Phone")
        self._email_var, _ = labeled_entry(f, "Email", width=40)
        self._fax_var,   _ = labeled_entry(f, "Fax Number")

        # ── Account Options ──────────────────────────────────────
        SectionLabel(f, "Account Options").pack(**pad)
        opts_frame = tk.Frame(f, bg=SURFACE2, padx=16, pady=12,
                              relief="solid", bd=1)
        opts_frame.pack(fill="x", padx=16, pady=(0, 8))

        self._acct_enabled_var   = tk.BooleanVar(value=True)
        self._must_change_pwd_var = tk.BooleanVar(value=True)
        self._pwd_never_expires_var = tk.BooleanVar(value=False)
        self._acct_disabled_var  = tk.BooleanVar(value=False)

        # Row 1: Enable / Disable toggle
        row1 = tk.Frame(opts_frame, bg=SURFACE2)
        row1.pack(fill="x", pady=3)
        cb_enabled = tk.Checkbutton(
            row1, text="Account is enabled",
            variable=self._acct_enabled_var, bg=SURFACE2,
            font=FONT_BODY, fg=TEXT, activebackground=SURFACE2,
            selectcolor=SURFACE2, cursor="hand2")
        cb_enabled.pack(side="left", padx=(0, 20))
        tk.Label(row1, text="—  uncheck to create account as DISABLED",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE2).pack(side="left")

        # Row 2: Must change password
        row2 = tk.Frame(opts_frame, bg=SURFACE2)
        row2.pack(fill="x", pady=3)
        cb_chpwd = tk.Checkbutton(
            row2, text="User must change password at next logon",
            variable=self._must_change_pwd_var, bg=SURFACE2,
            font=FONT_BODY, fg=TEXT, activebackground=SURFACE2,
            selectcolor=SURFACE2, cursor="hand2")
        cb_chpwd.pack(side="left", padx=(0, 20))
        tk.Label(row2, text="—  sets pwdLastSet=0 after password is applied",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE2).pack(side="left")

        # Row 3: Password never expires
        row3 = tk.Frame(opts_frame, bg=SURFACE2)
        row3.pack(fill="x", pady=3)
        cb_noexp = tk.Checkbutton(
            row3, text="Password never expires",
            variable=self._pwd_never_expires_var, bg=SURFACE2,
            font=FONT_BODY, fg=TEXT, activebackground=SURFACE2,
            selectcolor=SURFACE2, cursor="hand2")
        cb_noexp.pack(side="left", padx=(0, 20))
        tk.Label(row3, text="—  sets DONT_EXPIRE_PASSWORD flag (UAC bit)",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE2).pack(side="left")

        # Mutual exclusivity: "must change" and "never expires" can't both be set
        def _on_must_change(*_):
            if self._must_change_pwd_var.get():
                self._pwd_never_expires_var.set(False)
        def _on_never_exp(*_):
            if self._pwd_never_expires_var.get():
                self._must_change_pwd_var.set(False)
        self._must_change_pwd_var.trace_add("write", _on_must_change)
        self._pwd_never_expires_var.trace_add("write", _on_never_exp)

        # Summary label that updates dynamically
        self._opts_summary = tk.Label(opts_frame,
            text="", font=FONT_TINY, fg=ACCENT, bg=SURFACE2, anchor="w")
        self._opts_summary.pack(fill="x", pady=(6, 0))

        def _update_summary(*_):
            parts = []
            parts.append("Enabled" if self._acct_enabled_var.get() else "DISABLED")
            if self._must_change_pwd_var.get(): parts.append("Must change pwd at logon")
            if self._pwd_never_expires_var.get(): parts.append("Pwd never expires")
            self._opts_summary.config(text="Account will be created as:  " + "  ·  ".join(parts))
        self._acct_enabled_var.trace_add("write", _update_summary)
        self._must_change_pwd_var.trace_add("write", _update_summary)
        self._pwd_never_expires_var.trace_add("write", _update_summary)
        _update_summary()

        # ── Group Membership ─────────────────────────────────────
        SectionLabel(f, "Group Membership").pack(**pad)
        grp_frame = tk.Frame(f, bg=SURFACE, padx=16)
        grp_frame.pack(fill="x", pady=4)
        self._grp_selector = MultiSelectListbox(
            grp_frame,
            items=self._groups,
            label="Select groups to add this user to (hold Ctrl/Cmd for multiple):",
            height=6
        )
        self._grp_selector.pack(fill="x")
        tk.Button(grp_frame, text="↻ Refresh Groups", font=FONT_TINY,
                  bg=LIGHT_BLU, fg=ACCENT, relief="flat", padx=8, pady=3,
                  cursor="hand2", command=self.refresh_groups).pack(anchor="e", pady=(4, 0))

        # ── Buttons ──────────────────────────────────────────────
        btn_row = tk.Frame(f, bg=SURFACE, padx=16, pady=12)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="✓  Create User Account", font=FONT_SUB,
                  bg=ACCENT, fg=WHITE, relief="flat", padx=20, pady=8,
                  cursor="hand2", command=self._submit).pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="↺  Reset Form", font=FONT_BODY,
                  bg=BORDER, fg=TEXT, relief="flat", padx=12, pady=8,
                  cursor="hand2", command=self._reset).pack(side="left")
        self._result_lbl = tk.Label(f, text="", font=FONT_BODY,
                                    fg=GREEN, bg=SURFACE, wraplength=500)
        self._result_lbl.pack(fill="x", padx=16, pady=(0, 12))


    # ── Manager search helpers ────────────────────────────────────
    def _mgr_populate_users(self, users):
        """Called by refresh_users after AD fetch; populates manager dropdown."""
        self._mgr_all_users  = [(n, dn) for n, _, dn, _ in users]
        self._mgr_filtered   = list(self._mgr_all_users)

    def _mgr_filter(self, event=None):
        q = self._mgr_search_var.get().strip().lower()
        self._mgr_filtered = [(n, d) for n, d in self._mgr_all_users
                              if q in n.lower()] if q else list(self._mgr_all_users)
        self._mgr_show_popup(None)

    def _mgr_show_popup(self, event):
        items = self._mgr_filtered if self._mgr_filtered else self._mgr_all_users
        if not items:
            self._mgr_hide_popup()
            return
        self._mgr_hide_popup()
        win = tk.Toplevel(self)
        win.overrideredirect(True)
        win.lift()
        ex = self._mgr_entry.winfo_rootx()
        ey = self._mgr_entry.winfo_rooty() + self._mgr_entry.winfo_height()
        ew = self._mgr_entry.winfo_width() + 60
        height = min(180, len(items) * 22 + 6)
        win.geometry(f"{ew}x{height}+{ex}+{ey}")
        fr = tk.Frame(win, bg=SURFACE, bd=1, relief="solid")
        fr.pack(fill="both", expand=True)
        lb = tk.Listbox(fr, font=FONT_BODY, bd=0, relief="flat",
                        selectbackground=ACCENT, selectforeground=SURFACE,
                        activestyle="none", bg=SURFACE2, fg=TEXT,
                        exportselection=False)
        sb = ttk.Scrollbar(fr, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        for name, _ in items:
            lb.insert("end", name)
        lb.bind("<<ListboxSelect>>", lambda e: self._mgr_select(lb, items))
        self._mgr_popup_win = win
        self._mgr_popup_lb  = lb

    def _mgr_select(self, lb, items):
        sel = lb.curselection()
        if not sel:
            return
        name, dn = items[sel[0]]
        self._mgr_dn = dn
        self._mgr_search_var.set(name)
        self._mgr_entry.config(fg=TEXT)
        self._mgr_info.config(text=f"✓  {dn.split(',')[0].replace('CN=','')}", fg=TEAL)
        self._mgr_hide_popup()

    def _mgr_hide_popup(self):
        if self._mgr_popup_win and self._mgr_popup_win.winfo_exists():
            self._mgr_popup_win.destroy()
        self._mgr_popup_win = None

    def _mgr_clear(self):
        self._mgr_dn = ""
        self._mgr_search_var.set("")
        self._mgr_entry.config(fg=TEXT3)
        self._mgr_info.config(text="No manager selected", fg=TEXT3)

    def _validate(self):
        errors = []
        if not self._display_var.get().strip(): errors.append("Display Name is required.")
        if not self._first_var.get().strip():   errors.append("First Name is required.")
        if not self._last_var.get().strip():    errors.append("Last Name is required.")
        if not self._upn_var.get().strip():     errors.append("User Principal Name is required.")
        email = self._email_var.get().strip()
        if email and "@" not in email:
            errors.append("Email format invalid (must contain @).")
        return errors

    def _submit(self):
        errors = self._validate()
        if errors:
            messagebox.showerror("Validation Error", "\n".join(errors))
            return
        fields = {
            "displayName"                : self._display_var.get().strip(),
                        "givenName"                  : self._first_var.get().strip(),
            "sn"                         : self._last_var.get().strip(),
            "userPrincipalName"          : self._upn_var.get().strip(),
            "sAMAccountName"             : self._sam_var.get().strip(),
            "title"                      : self._title_var.get().strip(),
            "company"                    : self._company_var.get().strip(),
            "department"                 : self._dept_var.get().strip(),
            "physicalDeliveryOfficeName" : self._office_var.get().strip(),
            "manager"                    : self._mgr_dn if self._mgr_dn else self._mgr_search_var.get().strip(),
            "streetAddress"              : self._street_var.get().strip(),
            "l"                          : self._city_var.get().strip(),
            "st"                         : self._state_var.get().strip(),
            "postalCode"                 : self._zip_var.get().strip(),
            "co"                         : self._country_var.get().strip(),
            "telephoneNumber"            : self._phone_var.get().strip(),
            "mail"                       : self._email_var.get().strip(),
            "facsimileTelephoneNumber"   : self._fax_var.get().strip(),
        }
        selected_groups = self._grp_selector.get_selected_values()
        selected_labels = self._grp_selector.get_selected_labels()
        ou_dn = self._ou_dropdown.get_ou_dn()
        if not ou_dn:
            messagebox.showerror("OU Required",
                "Please select a Target OU before creating the user.\n\n"
                "Click '\u21bb Refresh OUs' to load available Organisational Units.")
            return
        ou = ou_dn

        # Determine if we have a real live AD connection
        real_ad = (self.ad.connected
                   and hasattr(self.ad, "conn")
                   and self.ad.conn is not None
                   and getattr(self.ad.conn, "bound", False))

        if real_ad:
            # Real AD
            try:
                password          = self._pwd_var.get() if self._pwd_var.get().strip() else None
                account_enabled   = self._acct_enabled_var.get()
                must_change_pwd   = self._must_change_pwd_var.get()
                pwd_never_exp     = self._pwd_never_expires_var.get()

                # AD hard rule: cannot enable an account without a password.
                # Intercept here and give the user a clear choice instead of
                # silently creating a disabled account.
                if account_enabled and not password:
                    answer = messagebox.askyesnocancel(
                        "Password Required to Enable Account",
                        "AD cannot enable an account without a password.\n\n"
                        "You have two options:\n\n"
                        "  YES  — Create the account as DISABLED\n"
                        "         (set a password and enable it manually in ADUC)\n\n"
                        "  NO   — Go back and set a password\n"
                        "         (use the Auto-generate button for a strong password)\n\n"
                        "Click YES to create disabled, NO to go back.")
                    if answer is None or answer is False:
                        return   # user cancelled or chose to go back
                    # User chose YES — create disabled
                    account_enabled = False
                ok, msg = self.ad.create_user(
                    ou, fields, selected_groups,
                    password         = password,
                    account_enabled  = account_enabled,
                    must_change_pwd  = must_change_pwd,
                    pwd_never_expires= pwd_never_exp
                )
                if ok:
                    self._result_lbl.config(fg=GREEN, text="✓ " + msg)
                    has_warnings  = "Warnings" in msg or "DISABLED" in msg or "Password could not" in msg
                    pwd_succeeded = bool(self._pwd_var.get().strip()) and "Password could not" not in msg
                    enabled       = self._acct_enabled_var.get()
                    must_chg      = self._must_change_pwd_var.get()
                    no_exp        = self._pwd_never_expires_var.get()

                    if pwd_succeeded and enabled:
                        acct_state = "✓ Enabled"
                    elif pwd_succeeded and not enabled:
                        acct_state = "Disabled (admin choice)"
                    else:
                        acct_state = "⚠ DISABLED — password not set; enable manually in ADUC"

                    opts_str = []
                    if must_chg: opts_str.append("Must change password at next logon")
                    if no_exp:   opts_str.append("Password never expires")
                    if not opts_str: opts_str.append("Standard password policy")
                    mgr_str = self._mgr_search_var.get().strip() or "Not set"
                    grp_str = ", ".join(selected_labels) if selected_labels else "None"
                    detail_msg = (
                        f"User account created!\n\n"
                        f"Display Name   : {fields['displayName']}\n"
                        f"UPN            : {fields['userPrincipalName']}\n"
                        f"Logon Name     : {fields['sAMAccountName'] or 'auto-generated'}\n"
                        f"Account state  : {acct_state}\n"
                        f"Password policy: {' | '.join(opts_str)}\n"
                        f"Manager        : {mgr_str}\n"
                        f"Groups joined  : {grp_str}\n\n"
                        f"{'⚠  Warnings present — check Activity Log for details.' if has_warnings else '✓  All attributes applied successfully.'}"
                    )
                    log_kind = "warn" if has_warnings else "ok"
                    self.log.log(f"User created: {fields['displayName']} | UPN:{fields['userPrincipalName']} | State:{acct_state}", log_kind)
                    self.status.set(f"User '{fields['displayName']}' created — {acct_state}.", log_kind)
                    messagebox.showinfo("User Created", detail_msg)
                else:
                    self._result_lbl.config(fg=RED, text="✗ " + msg)
                    self.log.log(f"User creation FAILED: {msg}", "error")
                    self.status.set("User creation failed — see log.", "error")
                    messagebox.showerror("Creation Failed", f"Could not create user:\n\n{msg}\n\nCheck:\n• OU path is correct\n• Account has write permission\n• User does not already exist")
            except Exception as ex:
                err = str(ex)
                self._result_lbl.config(fg=RED, text="✗ Error: " + err)
                self.log.log(f"Exception during user creation: {err}", "error")
                self.status.set("User creation error — see log.", "error")
                messagebox.showerror("Error", f"An error occurred:\n\n{err}")
        else:
            # Demo mode — simulate success
            grp_str = ", ".join(selected_labels) if selected_labels else "None"
            acct_state = "Enabled" if self._acct_enabled_var.get() else "DISABLED"
            chpwd      = "Yes" if self._must_change_pwd_var.get() else "No"
            noexp      = "Yes" if self._pwd_never_expires_var.get() else "No"
            msg = (f"[DEMO] User account prepared:\n"
                   f"  Display Name       : {fields['displayName']}\n"
                   f"  UPN                : {fields['userPrincipalName']}\n"
                   f"  Email              : {fields['mail']}\n"
                   f"  Manager            : {fields.get('manager','—') or '—'}\n"
                   f"  Department         : {fields['department'] or '—'}\n"
                   f"  Groups             : {grp_str}\n"
                   f"  Account state      : {acct_state}\n"
                   f"  Must change pwd    : {chpwd}\n"
                   f"  Pwd never expires  : {noexp}\n"
                   f"  OU                 : {ou}")
            self._result_lbl.config(fg=TEAL, text="✓ Demo: User would be created in AD")
            self.log.log(f"[DEMO] User: {fields['displayName']} | UPN: {fields['userPrincipalName']} | Groups: {grp_str}", "ok")
            self.status.set(f"[Demo] User '{fields['displayName']}' prepared.", "ok")
            messagebox.showinfo("Demo — User Created", msg)

    def _reset(self):
        for var in [self._display_var, self._first_var, self._last_var,
                    self._upn_var, self._sam_var, self._title_var,
                    self._company_var, self._dept_var, self._office_var,
                    self._street_var, self._city_var,
                    self._state_var, self._zip_var, self._phone_var,
                    self._email_var, self._fax_var]:
            var.set("")
        self._ou_dropdown.reset()
        self._country_var.set("")
        self._pwd_var.set("")
        self._pwd_show.set(False)
        self._pwd_ent.config(show="*")
        self._mgr_clear()
        self._acct_enabled_var.set(True)
        self._must_change_pwd_var.set(True)
        self._pwd_never_expires_var.set(False)
        self._result_lbl.config(text="")



# ─────────────────────────────────────────────────────────────────────────────
#  MANAGE USER — Set Password / Enable / Disable / Unlock
# ─────────────────────────────────────────────────────────────────────────────
class ManageUserForm(tk.Frame):
    """
    Lets an admin find an existing AD user and:
      • Set / reset their password
      • Enable or disable the account
      • Unlock a locked-out account
    Works over plain LDAP (port 389) via the StartTLS cascade in set_password().
    """

    def __init__(self, parent, ad: ADConnector, logger: ActivityLog, status: StatusBar):
        super().__init__(parent, bg=SURFACE)
        self.ad      = ad
        self.log     = logger
        self.status  = status
        self._users  = []        # (display_label, dn, sam, uac)
        self._sel_dn = ""
        self._sel_uac= 0
        self._build()

    # ── data refresh ─────────────────────────────────────────────
    def refresh_users(self):
        if not self.ad.connected:
            return
        raw = self.ad.get_all_users()
        self._users = []
        if raw:
            # Fetch UAC for each user in a second pass
            self._users = self._fetch_with_uac(raw)
        self._populate_list(self._users)
        self.log.log(f"Manage Users: {len(self._users)} user(s) loaded.", "ok")

    def _fetch_with_uac(self, raw_users):
        """Enrich user list with userAccountControl value."""
        enriched = []
        try:
            from ldap3 import SUBTREE
            self.ad.conn.search(
                search_base   = self.ad.base_dn,
                search_filter = "(&(objectClass=user)(objectCategory=person))",
                search_scope  = SUBTREE,
                attributes    = ["sAMAccountName", "distinguishedName",
                                 "displayName", "userAccountControl",
                                 "lockoutTime"],
                size_limit    = 1000
            )
            uac_map = {}
            lock_map = {}
            for e in self.ad.conn.entries:
                try:
                    sam = str(e.sAMAccountName).lower()
                    uac = int(str(e.userAccountControl)) if e.userAccountControl else 512
                    lock= int(str(e.lockoutTime))        if e.lockoutTime        else 0
                    uac_map[sam]  = uac
                    lock_map[sam] = lock
                except Exception:
                    pass
            for name, sam, dn, mail in raw_users:
                uac  = uac_map.get(sam.lower(), 512)
                lock = lock_map.get(sam.lower(), 0)
                enriched.append((name, sam, dn, mail, uac, lock))
        except Exception:
            for name, sam, dn, mail in raw_users:
                enriched.append((name, sam, dn, mail, 512, 0))
        return enriched

    def _populate_list(self, users):
        self._lb.delete(0, "end")
        for name, sam, dn, mail, uac, lock in users:
            disabled = bool(uac & 2)
            locked   = lock > 0
            tag = " 🔒 LOCKED" if locked else (" ⛔ DISABLED" if disabled else " ✓ Enabled")
            self._lb.insert("end", f"{name}  ({sam}){tag}")
        self._filtered = users

    # ── UI build ─────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg="#FFF7ED", padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔑  Manage User Accounts",
                 font=FONT_HEAD, fg=NAVY, bg="#FFF7ED").pack(side="left")
        tk.Label(hdr, text="Select a user to edit attributes, password, account state or delete",
                 font=FONT_TINY, fg=TEXT2, bg="#FFF7ED").pack(side="right")

        # Split pane
        paned = tk.PanedWindow(self, orient="horizontal", bg=BG,
                               sashwidth=6, relief="flat")
        paned.pack(fill="both", expand=True)

        # ── LEFT: user list ───────────────────────────────────────
        left = tk.Frame(paned, bg=SURFACE, width=300)
        paned.add(left, minsize=250)

        srch_row = tk.Frame(left, bg=SURFACE)
        srch_row.pack(fill="x", padx=8, pady=6)
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._filter_users)
        tk.Entry(srch_row, textvariable=self._search_var, font=FONT_BODY,
                 relief="solid", bd=1, bg=SURFACE2,
                 highlightthickness=1, highlightcolor=ACCENT).pack(
                 side="left", fill="x", expand=True)
        tk.Button(srch_row, text="↻", font=FONT_BODY, bg=LIGHT_BLU, fg=ACCENT,
                  relief="flat", padx=6, cursor="hand2",
                  command=self.refresh_users).pack(side="right", padx=(4,0))

        lb_frame = tk.Frame(left, bg=SURFACE)
        lb_frame.pack(fill="both", expand=True, padx=8, pady=(0,4))
        self._lb = tk.Listbox(lb_frame, font=FONT_BODY, bd=1, relief="solid",
                              exportselection=False, selectbackground=ACCENT,
                              selectforeground=WHITE, activestyle="none",
                              bg=SURFACE2, fg=TEXT)
        sb = ttk.Scrollbar(lb_frame, orient="vertical", command=self._lb.yview)
        self._lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._lb.pack(side="left", fill="both", expand=True)
        self._lb.bind("<<ListboxSelect>>", self._on_select)
        self._filtered = []

        # ── RIGHT: tabbed edit area ───────────────────────────────
        right = tk.Frame(paned, bg=SURFACE)
        paned.add(right, minsize=500)

        # Placeholder when no user selected
        self._no_sel_lbl = tk.Label(right,
            text="← Select a user from the list to manage",
            font=FONT_BODY, fg=TEXT3, bg=SURFACE)
        self._no_sel_lbl.place(relx=0.5, rely=0.35, anchor="center")

        # Selected user info bar
        self._sel_info = tk.Label(right, text="",
                                  font=FONT_TINY, fg=TEAL, bg=LIGHT_BLU,
                                  anchor="w", padx=10, pady=5)
        # Tab bar
        self._tab_frame = tk.Frame(right, bg=SURFACE)
        self._tab_keys  = ["attrs", "password", "state", "delete"]
        self._tab_labels= {"attrs": "✏ Edit Attributes",
                           "password": "🔑 Password",
                           "state": "⚡ Account State",
                           "delete": "🗑 Delete"}
        self._tab_btns  = {}
        tab_bar = tk.Frame(right, bg=LIGHT_BLU)
        for key in self._tab_keys:
            b = tk.Button(tab_bar, text=self._tab_labels[key],
                          font=FONT_TINY, relief="flat", padx=12, pady=6,
                          cursor="hand2",
                          command=lambda k=key: self._show_tab(k))
            b.pack(side="left", padx=2, pady=4)
            self._tab_btns[key] = b

        # Tab content frames (scrollable)
        self._tab_contents = {}
        for key in self._tab_keys:
            canvas = tk.Canvas(right, bg=SURFACE, highlightthickness=0)
            tsb = ttk.Scrollbar(right, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=tsb.set)
            inner = tk.Frame(canvas, bg=SURFACE)
            wid = canvas.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>", lambda e, c=canvas: c.configure(
                scrollregion=c.bbox("all")))
            canvas.bind("<Configure>", lambda e, c=canvas, w=wid: c.itemconfig(w, width=e.width))
            _bind_mousewheel(canvas)
            self._tab_contents[key] = (canvas, tsb, inner)

        # ── Tab: Edit Attributes ──────────────────────────────────
        f = self._tab_contents["attrs"][2]
        SectionLabel(f, "Identity").pack(fill="x", padx=12, pady=(8,2))
        self._edit_vars = {}
        all_edit_fields = [
            # (label, attr, section_after)
            ("Display Name",         "displayName",                  None),
            ("First Name",           "givenName",                    None),
            ("Last Name",            "sn",                           None),
            ("User Principal Name",  "userPrincipalName",            None),
            ("Logon Name (SAM)",     "sAMAccountName",               "Job Information"),
            ("Job Title",            "title",                        None),
            ("Department",           "department",                   None),
            ("Company",              "company",                      None),
            ("Office Location",      "physicalDeliveryOfficeName",   None),
            # manager handled separately below as searchable dropdown
            ("Email",                "mail",                         None),
            ("Business Phone",       "telephoneNumber",              None),
            ("Fax",                  "facsimileTelephoneNumber",     "Address"),
            ("Street Address",       "streetAddress",                None),
            ("City",                 "l",                            None),
            ("State / Province",     "st",                           None),
            ("Postal Code",          "postalCode",                   None),
            ("Country",              "co",                           None),
        ]
        for label, attr, section in all_edit_fields:
            if section:
                SectionLabel(f, section).pack(fill="x", padx=12, pady=(8,2))
            var, _ = labeled_entry(f, label, width=36)
            self._edit_vars[attr] = var

        # ── Manager searchable dropdown (Edit Attributes) ──────────
        SectionLabel(f, "Contact").pack(fill="x", padx=12, pady=(8,2))
        mgr_edit_outer = tk.Frame(f, bg=SURFACE)
        mgr_edit_outer.pack(fill="x", pady=3, padx=8)
        tk.Label(mgr_edit_outer, text="Manager", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE, width=22, anchor="w").pack(side="left")
        mgr_edit_inner = tk.Frame(mgr_edit_outer, bg=SURFACE)
        mgr_edit_inner.pack(side="left", fill="x", expand=True)

        self._edit_mgr_search_var = tk.StringVar()
        self._edit_mgr_dn         = ""
        self._edit_mgr_all_users  = []
        self._edit_mgr_filtered   = []
        self._edit_mgr_popup_win  = None

        self._edit_mgr_entry = tk.Entry(mgr_edit_inner, textvariable=self._edit_mgr_search_var,
                                        font=FONT_BODY, width=34, relief="solid", bd=1,
                                        highlightthickness=1, highlightcolor=ACCENT, bg=SURFACE2)
        self._edit_mgr_entry.pack(side="left")
        self._edit_mgr_entry.bind("<KeyRelease>", self._edit_mgr_filter)
        self._edit_mgr_entry.bind("<FocusOut>",   lambda e: self.after(200, self._edit_mgr_hide_popup))

        tk.Button(mgr_edit_inner, text="✕", font=FONT_TINY, bg=SURFACE, fg=TEXT3,
                  relief="flat", padx=4, cursor="hand2",
                  command=self._edit_mgr_clear).pack(side="left", padx=(4,0))

        self._edit_mgr_info = tk.Label(mgr_edit_inner, text="Type to search manager",
                                       font=FONT_TINY, fg=TEXT3, bg=SURFACE)
        self._edit_mgr_info.pack(side="left", padx=(8,0))

        btn_row = tk.Frame(f, bg=SURFACE, padx=12, pady=10)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="💾  Save Changes", font=FONT_SUB,
                  bg=ACCENT, fg=WHITE, relief="flat", padx=16, pady=7,
                  cursor="hand2", command=self._do_edit_user).pack(side="left", padx=(0,8))
        self._edit_result = tk.Label(f, text="", font=FONT_BODY,
                                     fg=GREEN, bg=SURFACE, wraplength=460)
        self._edit_result.pack(fill="x", padx=12, pady=(0,12))

        # ── Tab: Password ─────────────────────────────────────────
        fp = self._tab_contents["password"][2]
        SectionLabel(fp, "Set / Reset Password").pack(fill="x", padx=12, pady=(8,4))
        pwd_frame = tk.Frame(fp, bg=SURFACE2, padx=14, pady=12, relief="solid", bd=1)
        pwd_frame.pack(fill="x", padx=12, pady=(0,12))

        pr1 = tk.Frame(pwd_frame, bg=SURFACE2)
        pr1.pack(fill="x", pady=(0,6))
        tk.Label(pr1, text="New Password", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE2, width=16, anchor="w").pack(side="left")
        self._new_pwd_var  = tk.StringVar()
        self._pwd_show_var = tk.BooleanVar(value=False)
        self._new_pwd_ent  = tk.Entry(pr1, textvariable=self._new_pwd_var,
                                      show="*", font=FONT_BODY, width=24,
                                      relief="solid", bd=1, bg=SURFACE)
        self._new_pwd_ent.pack(side="left", padx=(4,8))
        def _toggle():
            self._new_pwd_ent.config(show="" if self._pwd_show_var.get() else "*")
        tk.Checkbutton(pr1, text="Show", variable=self._pwd_show_var,
                       font=FONT_TINY, bg=SURFACE2, cursor="hand2",
                       command=_toggle).pack(side="left", padx=(0,8))
        import string, secrets as _sec
        def _gen():
            alpha = string.ascii_letters + string.digits + "!@#$%^&*"
            while True:
                p = "".join(_sec.choice(alpha) for _ in range(12))
                if (any(c.isupper() for c in p) and any(c.islower() for c in p)
                        and any(c.isdigit() for c in p)
                        and any(c in "!@#$%^&*" for c in p)):
                    break
            self._new_pwd_var.set(p)
            self._pwd_show_var.set(True)
            self._new_pwd_ent.config(show="")
        tk.Button(pr1, text="⚙ Auto-generate", font=FONT_TINY,
                  bg=ACCENT, fg=WHITE, relief="flat", padx=8, pady=3,
                  cursor="hand2", command=_gen).pack(side="left")

        pr2 = tk.Frame(pwd_frame, bg=SURFACE2)
        pr2.pack(fill="x", pady=(0,6))
        self._must_chg_var = tk.BooleanVar(value=True)
        tk.Checkbutton(pr2, text="User must change password at next logon",
                       variable=self._must_chg_var, font=FONT_BODY,
                       fg=TEXT, bg=SURFACE2, activebackground=SURFACE2,
                       cursor="hand2").pack(side="left")
        tk.Button(pwd_frame, text="🔑  Set Password & Enable Account", font=FONT_SUB,
                  bg=ACCENT, fg=WHITE, relief="flat", padx=16, pady=6,
                  cursor="hand2", command=self._do_set_password).pack(anchor="w", pady=(4,0))
        self._pwd_result = tk.Label(fp, text="", font=FONT_BODY,
                                    fg=GREEN, bg=SURFACE, wraplength=460)
        self._pwd_result.pack(fill="x", padx=12, pady=(4,0))

        # ── Tab: Account State ────────────────────────────────────
        fs = self._tab_contents["state"][2]
        SectionLabel(fs, "Enable / Disable Account").pack(fill="x", padx=12, pady=(8,4))
        state_frame = tk.Frame(fs, bg=SURFACE2, padx=14, pady=12, relief="solid", bd=1)
        state_frame.pack(fill="x", padx=12, pady=(0,12))
        en_row = tk.Frame(state_frame, bg=SURFACE2)
        en_row.pack(fill="x")
        tk.Button(en_row, text="✓  Enable Account", font=FONT_BODY,
                  bg=GREEN, fg=WHITE, relief="flat", padx=14, pady=6,
                  cursor="hand2",
                  command=lambda: self._do_set_uac(enable=True)).pack(side="left", padx=(0,8))
        tk.Button(en_row, text="⛔  Disable Account", font=FONT_BODY,
                  bg=RED, fg=WHITE, relief="flat", padx=14, pady=6,
                  cursor="hand2",
                  command=lambda: self._do_set_uac(enable=False)).pack(side="left")
        tk.Label(state_frame, text="Toggles the account enabled/disabled state immediately.",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE2).pack(anchor="w", pady=(8,0))

        SectionLabel(fs, "Unlock Account").pack(fill="x", padx=12, pady=(8,4))
        lock_frame = tk.Frame(fs, bg=SURFACE2, padx=14, pady=12, relief="solid", bd=1)
        lock_frame.pack(fill="x", padx=12, pady=(0,12))
        tk.Button(lock_frame, text="🔓  Unlock Account", font=FONT_BODY,
                  bg=TEAL, fg=WHITE, relief="flat", padx=14, pady=6,
                  cursor="hand2", command=self._do_unlock).pack(anchor="w")
        tk.Label(lock_frame,
                 text="Clears a lockout (lockoutTime=0) without changing the password.",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE2).pack(anchor="w", pady=(8,0))

        self._state_result = tk.Label(fs, text="", font=FONT_BODY,
                                      fg=GREEN, bg=SURFACE, wraplength=460)
        self._state_result.pack(fill="x", padx=12, pady=(0,12))

        # ── Tab: Delete ───────────────────────────────────────────
        fd = self._tab_contents["delete"][2]
        SectionLabel(fd, "Delete User Account").pack(fill="x", padx=12, pady=(8,4))
        del_frame = tk.Frame(fd, bg="#FEF2F2", padx=14, pady=16, relief="solid", bd=1)
        del_frame.pack(fill="x", padx=12, pady=(0,12))
        tk.Label(del_frame,
                 text="⚠  This permanently removes the user account from Active Directory.\n"
                      "All associated data will be lost. This action cannot be undone.",
                 font=FONT_BODY, fg=RED, bg="#FEF2F2", justify="left").pack(anchor="w", pady=(0,12))
        tk.Button(del_frame, text="🗑  Permanently Delete Selected User", font=FONT_SUB,
                  bg=RED, fg=WHITE, relief="flat", padx=16, pady=7,
                  cursor="hand2", command=self._do_delete_user).pack(anchor="w")
        self._del_result = tk.Label(fd, text="", font=FONT_BODY,
                                    fg=AMBER, bg=SURFACE, wraplength=460)
        self._del_result.pack(fill="x", padx=12, pady=(4,0))

        # Wire tab_bar and content; initially hidden until user selected
        self._tab_bar_widget = tab_bar
        self._right_frame = right
        self._cur_tab = "attrs"

    def _show_tab(self, key):
        self._cur_tab = key
        # Hide all tab content
        for k, (canvas, tsb, _) in self._tab_contents.items():
            canvas.place_forget()
            tsb.place_forget()
        # Show selected
        canvas, tsb, _ = self._tab_contents[key]
        canvas.place(x=0, rely=0.12, relwidth=0.95, relheight=0.88)
        tsb.place(relx=0.95, rely=0.12, relwidth=0.05, relheight=0.88)
        # Update tab button colours
        colors_map = {"attrs": ACCENT, "password": "#7C3AED",
                      "state": TEAL, "delete": RED}
        for k, b in self._tab_btns.items():
            b.config(bg=colors_map[k] if k==key else BORDER,
                     fg=WHITE if k==key else TEXT)

    # ── interaction helpers ───────────────────────────────────────
    def _on_select(self, event=None):
        sel = self._lb.curselection()
        if not sel or sel[0] >= len(self._filtered):
            return
        u = self._filtered[sel[0]]
        name, sam, dn, mail, uac, lock = u
        self._sel_dn   = dn
        self._sel_uac  = uac
        self._sel_name = name
        disabled = bool(uac & 2)
        locked   = lock > 0
        state    = ("🔒 LOCKED-OUT" if locked
                    else ("⛔ DISABLED" if disabled else "✓ Enabled"))

        # Show the info bar and tab bar now a user is selected
        self._no_sel_lbl.place_forget()
        self._sel_info.config(
            text=f"  {name}  |  {sam}  |  {state}  |  {dn}",
            fg=RED if (disabled or locked) else TEAL, bg=LIGHT_BLU)
        self._sel_info.place(relx=0, rely=0, relwidth=1)
        self._tab_bar_widget.place(relx=0, rely=0.06, relwidth=1)
        self._show_tab(self._cur_tab)

        # Auto-load current AD attributes into edit fields
        if self.ad.connected and getattr(self.ad.conn, "bound", False):
            attrs = self.ad.get_user_attributes(dn)
            for attr, var in self._edit_vars.items():
                val = attrs.get(attr, "")
                var.set(val)
            # Populate manager dropdown from all users
            raw_users = self.ad.get_all_users()
            self._edit_mgr_all_users = [(n, d) for n, _, d, _ in raw_users]
            self._edit_mgr_filtered  = list(self._edit_mgr_all_users)
            # Pre-fill manager field from loaded attrs
            mgr_raw = attrs.get("manager", "")
            if mgr_raw and "CN=" in str(mgr_raw):
                mgr_cn = str(mgr_raw).split(",")[0].replace("CN=","").replace("cn=","")
                self._edit_mgr_dn = str(mgr_raw)
                self._edit_mgr_search_var.set(mgr_cn)
                self._edit_mgr_info.config(text=f"✓  {mgr_cn}", fg=TEAL)
            else:
                self._edit_mgr_dn = ""
                self._edit_mgr_search_var.set("")
                self._edit_mgr_info.config(text="Type to search manager", fg=TEXT3)
            self._edit_result.config(text="✓ Current values loaded — edit and Save Changes.", fg=TEAL)

    # ── Edit-tab manager search helpers ──────────────────────────
    def _edit_mgr_filter(self, event=None):
        q = self._edit_mgr_search_var.get().strip().lower()
        self._edit_mgr_filtered = [(n, d) for n, d in self._edit_mgr_all_users
                                   if q in n.lower()] if q else list(self._edit_mgr_all_users)
        self._edit_mgr_show_popup()

    def _edit_mgr_show_popup(self):
        items = self._edit_mgr_filtered if self._edit_mgr_filtered else self._edit_mgr_all_users
        if not items:
            self._edit_mgr_hide_popup()
            return
        self._edit_mgr_hide_popup()
        win = tk.Toplevel(self)
        win.overrideredirect(True)
        win.lift()
        ex = self._edit_mgr_entry.winfo_rootx()
        ey = self._edit_mgr_entry.winfo_rooty() + self._edit_mgr_entry.winfo_height()
        ew = self._edit_mgr_entry.winfo_width() + 60
        height = min(180, len(items) * 22 + 6)
        win.geometry(f"{ew}x{height}+{ex}+{ey}")
        fr = tk.Frame(win, bg=SURFACE, bd=1, relief="solid")
        fr.pack(fill="both", expand=True)
        lb = tk.Listbox(fr, font=FONT_BODY, bd=0, relief="flat",
                        selectbackground=ACCENT, selectforeground=SURFACE,
                        activestyle="none", bg=SURFACE2, fg=TEXT,
                        exportselection=False)
        sb = ttk.Scrollbar(fr, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        for name, _ in items:
            lb.insert("end", name)
        lb.bind("<<ListboxSelect>>", lambda e: self._edit_mgr_select(lb, items))
        self._edit_mgr_popup_win = win

    def _edit_mgr_select(self, lb, items):
        sel = lb.curselection()
        if not sel:
            return
        name, dn = items[sel[0]]
        self._edit_mgr_dn = dn
        self._edit_mgr_search_var.set(name)
        self._edit_mgr_entry.config(fg=TEXT)
        self._edit_mgr_info.config(text=f"✓  {dn.split(chr(44))[0].replace(chr(67)+chr(78)+chr(61),chr(32)).strip()}", fg=TEAL)
        self._edit_mgr_hide_popup()

    def _edit_mgr_hide_popup(self):
        if self._edit_mgr_popup_win and self._edit_mgr_popup_win.winfo_exists():
            self._edit_mgr_popup_win.destroy()
        self._edit_mgr_popup_win = None

    def _edit_mgr_clear(self):
        self._edit_mgr_dn = ""
        self._edit_mgr_search_var.set("")
        self._edit_mgr_info.config(text="Type to search manager", fg=TEXT3)

    def _filter_users(self, *_):
        q = self._search_var.get().strip().lower()
        filtered = [u for u in self._users
                    if q in u[0].lower() or q in u[1].lower()] if q else list(self._users)
        self._populate_list(filtered)

    def _require_selection(self):
        if not self._sel_dn:
            messagebox.showwarning("No User Selected",
                "Please select a user from the list above first.")
            return False
        return True

    def _require_real_ad(self):
        real = (self.ad.connected and hasattr(self.ad, "conn")
                and self.ad.conn is not None
                and getattr(self.ad.conn, "bound", False))
        if not real:
            messagebox.showwarning("Not Connected",
                "This action requires a live AD connection.\n"
                "Connect via File → Connect to AD first.")
        return real

    def _do_edit_user(self):
        if not self._require_selection() or not self._require_real_ad():
            return
        changes = {attr: var.get().strip()
                   for attr, var in self._edit_vars.items()}
        # Inject manager DN from the searchable dropdown
        mgr_dn = getattr(self, "_edit_mgr_dn", "").strip()
        if mgr_dn:
            changes["manager"] = mgr_dn
        elif getattr(self, "_edit_mgr_search_var", None) and self._edit_mgr_search_var.get().strip():
            # User typed a name but did not pick from list — resolve via AD
            resolved = self.ad.get_manager_dn(self._edit_mgr_search_var.get().strip())
            if resolved:
                changes["manager"] = resolved
        # Remove empty fields (clear them in AD)
        changes = {k: v for k, v in changes.items() if v != ""}
        if not changes:
            messagebox.showinfo("No Changes", "Load current values first, then edit fields to change.")
            return
        ok, msg = self.ad.modify_user(self._sel_dn, changes)
        name = getattr(self, "_sel_name", "user")
        if ok:
            _record("modif", "user", name)
            self._edit_result.config(fg=GREEN, text=f"✓ {name} updated successfully.")
            self.log.log(f"User modified: {name} | Fields: {list(changes.keys())}", "ok")
            self.status.set(f"User '{name}' updated.", "ok")
            self.after(600, self.refresh_users)
        else:
            self._edit_result.config(fg=RED, text=f"✗ Update failed: {msg}")
            self.log.log(f"User modify FAILED: {msg}", "error")

    def _do_delete_user(self):
        if not self._require_selection() or not self._require_real_ad():
            return
        name = getattr(self, "_sel_name", self._sel_dn)
        if not messagebox.askyesno("Confirm Delete",
                f"Permanently delete user:\n\n  {name}\n\n"
                f"DN: {self._sel_dn}\n\n"
                f"⚠ This cannot be undone.", icon="warning"):
            return
        ok, msg = self.ad.delete_user(self._sel_dn)
        if ok:
            _record("delet", "user", name)
            self._del_result.config(fg=AMBER, text=f"✓ User '{name}' deleted.")
            self.log.log(f"User DELETED: {name}", "warn")
            self.status.set(f"User '{name}' deleted.", "warn")
            self._sel_dn = ""
            self._sel_name = ""
            self._sel_info.config(text="No user selected", fg=TEXT2, bg=LIGHT_BLU)
            for v in self._edit_vars.values():
                v.set("")
            self.after(400, self.refresh_users)
        else:
            self._del_result.config(fg=RED, text=f"✗ Delete failed: {msg}")
            self.log.log(f"User delete FAILED: {msg}", "error")
            messagebox.showerror("Delete Failed", msg)

    # ── actions ───────────────────────────────────────────────────
    def _do_set_password(self):
        if not self._require_selection() or not self._require_real_ad():
            return
        pwd = self._new_pwd_var.get().strip()
        if not pwd:
            messagebox.showerror("Password Required",
                "Please enter a new password or use Auto-generate.")
            return

        ok, method, err = self.ad.set_password(self._sel_dn, pwd)
        name = getattr(self, "_sel_name", "selected user")

        if ok:
            self._pwd_result.config(fg=GREEN,
                text=f"✓ Password set for {name} via {method}.")
            self.log.log(f"Password set: {name} | Method: {method}", "ok")
            self.status.set(f"Password set for {name}.", "ok")

            # Must-change-password flag
            if self._must_chg_var.get():
                self.ad.conn.modify(self._sel_dn,
                    {"pwdLastSet": [(MODIFY_REPLACE, [0])]})

            # Auto-enable if account was disabled
            if self._sel_uac & 2:
                uac_new = self._sel_uac & ~2   # clear ACCOUNTDISABLE bit
                uac_ok  = self.ad.conn.modify(self._sel_dn,
                    {"userAccountControl": [(MODIFY_REPLACE, [uac_new])]})
                if uac_ok:
                    self._pwd_result.config(fg=GREEN,
                        text=f"✓ Password set AND account enabled for {name} via {method}.")
                    self.log.log(f"Account enabled after password set: {name}", "ok")
                    self.status.set(f"Password set & account enabled for {name}.", "ok")
                else:
                    self._pwd_result.config(fg=AMBER,
                        text=f"✓ Password set, but enabling failed: "
                             f"{self.ad.conn.result.get('description','')}")

            # Refresh list to show updated state
            self.after(500, self.refresh_users)
            messagebox.showinfo("Password Set",
                f"Password successfully set for:\n{name}\n\n"
                f"Method used: {method}\n"
                f"{'Account was disabled — it has been automatically enabled.' if self._sel_uac & 2 else ''}")
        else:
            self._pwd_result.config(fg=RED,
                text=f"✗ Failed to set password: {err}")
            self.log.log(f"Password set FAILED for {name}: {err}", "error")
            self.status.set("Password set failed — see log.", "error")
            messagebox.showerror("Password Failed",
                f"Could not set password for {name}.\n\n"
                f"All 4 methods were tried:\n"
                f"  1. unicodePwd on existing connection\n"
                f"  2. Extended Password Modify op\n"
                f"  3. Fresh LDAPS connection on port 636\n"
                f"  4. StartTLS upgrade on port 389\n\n"
                f"Last error: {err}\n\n"
                f"Most likely cause: port 636 is blocked by firewall on the DC.\n"
                f"To fix: open port 636 on the DC firewall, or ask your AD admin\n"
                f"to enable LDAPS / install a DC certificate.")

    def _do_set_uac(self, enable: bool):
        if not self._require_selection() or not self._require_real_ad():
            return
        name = getattr(self, "_sel_name", "selected user")
        uac  = self._sel_uac
        if enable:
            uac_new = uac & ~2          # clear ACCOUNTDISABLE bit
            verb    = "enable"
        else:
            uac_new = uac | 2           # set ACCOUNTDISABLE bit
            verb    = "disable"

        ok = self.ad.conn.modify(self._sel_dn,
            {"userAccountControl": [(MODIFY_REPLACE, [uac_new])]})
        if ok:
            self._sel_uac = uac_new
            state = "Enabled ✓" if enable else "Disabled ⛔"
            self._state_result.config(fg=GREEN if enable else RED,
                text=f"✓ Account {state} for {name}.")
            self.log.log(f"Account {verb}d: {name}", "ok")
            self.status.set(f"Account {verb}d: {name}.", "ok")
            self.after(500, self.refresh_users)
        else:
            err = self.ad.conn.result.get("description", "unknown")
            self._state_result.config(fg=RED,
                text=f"✗ Could not {verb} account: {err}")
            self.log.log(f"Account {verb} FAILED for {name}: {err}", "error")

    def _do_unlock(self):
        if not self._require_selection() or not self._require_real_ad():
            return
        name = getattr(self, "_sel_name", "selected user")
        # Clear lockoutTime by setting it to 0
        ok = self.ad.conn.modify(self._sel_dn,
            {"lockoutTime": [(MODIFY_REPLACE, [0])]})
        if ok:
            self._state_result.config(fg=GREEN,
                text=f"✓ Account unlocked for {name}.")
            self.log.log(f"Account unlocked: {name}", "ok")
            self.status.set(f"Account unlocked: {name}.", "ok")
            self.after(500, self.refresh_users)
        else:
            err = self.ad.conn.result.get("description", "unknown")
            self._state_result.config(fg=RED,
                text=f"✗ Unlock failed: {err}")
            self.log.log(f"Unlock FAILED for {name}: {err}", "error")


# ─────────────────────────────────────────────────────────────────────────────
#  GROUP CREATION FORM
# ─────────────────────────────────────────────────────────────────────────────
class GroupCreationForm(tk.Frame):
    GROUP_TYPES = [
        "Global Security",
        "Domain Local Security",
        "Universal Security",
        "Global Distribution",
        "Domain Local Dist.",
        "Universal Distribution",
    ]

    def __init__(self, parent, ad: ADConnector, logger: ActivityLog, status: StatusBar):
        super().__init__(parent, bg=SURFACE)
        self.ad     = ad
        self.log    = logger
        self.status = status
        self._users = []
        self._ous   = []
        self._build()

    def refresh_ous(self):
        if self.ad.connected:
            self._ous = self.ad.get_all_ous()
            self._ou_dropdown.update_ous(self._ous)
            count = len(self._ous)
            if count <= 1:
                err = getattr(self.ad, "_last_error", "")
                err_hint = f" Error: {err}" if err else " Check AD permissions — account needs 'Read' on OU objects."
                self.log.log(f"Group OUs loaded: {count} OU(s). Only root DN available.{err_hint}", "warn")
            else:
                self.log.log(f"Group OUs loaded: {count} OU(s) found.", "ok")

    def refresh_users(self):
        if self.ad.connected:
            raw = self.ad.get_all_users()
            self._users = [(f"{n} <{m}>" if m else n, dn) for n, _, dn, m in raw]
            self._user_selector.update_items(self._users)
            if self._users:
                self.log.log(f"Users for group membership loaded: {len(self._users)} user(s).", "ok")
            else:
                err = getattr(self.ad, "_last_error", "")
                err_hint = f" AD error: {err}" if err else ""
                self.log.log(f"No users found for group membership.{err_hint}", "warn")

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg="#E8F5E9", padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="👥  Create Group",
                 font=FONT_HEAD, fg=NAVY, bg="#E8F5E9").pack(side="left")
        tk.Label(hdr, text="Create AD Security or Distribution groups",
                 font=FONT_TINY, fg=TEXT2, bg="#E8F5E9").pack(side="right")

        # Scrollable body
        canvas = tk.Canvas(self, bg=SURFACE, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._inner = tk.Frame(canvas, bg=SURFACE)
        self._win   = canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._win, width=e.width))
        _bind_mousewheel(canvas)

        f = self._inner
        pad = {"padx": 16, "pady": 4, "fill": "x"}

        # ── OU ───────────────────────────────────────────────────
        SectionLabel(f, "Organisational Unit (OU)").pack(**pad)
        self._ou_dropdown = OUDropdown(f)
        self._ou_dropdown.pack(fill="x", padx=8, pady=4)
        self._ou_dropdown.set_refresh_command(self.refresh_ous)

        # ── Group Details ────────────────────────────────────────
        SectionLabel(f, "Group Details").pack(**pad)
        self._name_var,  _ = labeled_entry(f, "Group Name  *",  required=True)
        self._type_var,  _ = labeled_combo(f, "Group Type  *", self.GROUP_TYPES)
        self._desc_var,  _ = labeled_entry(f, "Description", width=48)

        # Group type explanation
        info_row = tk.Frame(f, bg=LIGHT_BLU, padx=12, pady=6)
        info_row.pack(fill="x", padx=16, pady=4)
        self._type_info = tk.Label(info_row,
            text="Global Security: Most common — use for resource access control within a domain.",
            font=FONT_TINY, fg=ACCENT, bg=LIGHT_BLU, wraplength=560, justify="left")
        self._type_info.pack(anchor="w")
        type_descs = {
            "Global Security"       : "Global Security: Most common — controls access to resources within a domain.",
            "Domain Local Security" : "Domain Local Security: Assign permissions to resources in the local domain.",
            "Universal Security"    : "Universal Security: Use across multiple domains in a forest.",
            "Global Distribution"   : "Global Distribution: Email distribution list with global scope.",
            "Domain Local Dist."    : "Domain Local Distribution: Email distribution list for the local domain.",
            "Universal Distribution": "Universal Distribution: Email distribution list across the entire forest.",
        }
        def _update_desc(*_):
            self._type_info.config(text=type_descs.get(self._type_var.get(), ""))
        self._type_var.trace_add("write", _update_desc)

        # ── Member Selection ─────────────────────────────────────
        SectionLabel(f, "Add Members").pack(**pad)
        usr_frame = tk.Frame(f, bg=SURFACE, padx=16)
        usr_frame.pack(fill="x", pady=4)
        self._user_selector = MultiSelectListbox(
            usr_frame,
            items=self._users,
            label="Select users to add to this group (hold Ctrl/Cmd for multiple):",
            height=8
        )
        self._user_selector.pack(fill="x")
        tk.Button(usr_frame, text="↻ Refresh Users", font=FONT_TINY,
                  bg=LIGHT_BLU, fg=ACCENT, relief="flat", padx=8, pady=3,
                  cursor="hand2", command=self.refresh_users).pack(anchor="e", pady=(4, 0))

        # ── Buttons ──────────────────────────────────────────────
        btn_row = tk.Frame(f, bg=SURFACE, padx=16, pady=12)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="✓  Create Group", font=FONT_SUB,
                  bg=TEAL, fg=WHITE, relief="flat", padx=20, pady=8,
                  cursor="hand2", command=self._submit).pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="↺  Reset Form", font=FONT_BODY,
                  bg=BORDER, fg=TEXT, relief="flat", padx=12, pady=8,
                  cursor="hand2", command=self._reset).pack(side="left")
        self._result_lbl = tk.Label(f, text="", font=FONT_BODY,
                                    fg=GREEN, bg=SURFACE, wraplength=500)
        self._result_lbl.pack(fill="x", padx=16, pady=(0, 12))

    def _submit(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showerror("Validation Error", "Group Name is required.")
            return
        gtype       = self._type_var.get()
        desc        = self._desc_var.get().strip()
        member_dns  = self._user_selector.get_selected_values()
        member_lbls = self._user_selector.get_selected_labels()
        ou_dn       = self._ou_dropdown.get_ou_dn()
        if not ou_dn:
            messagebox.showerror("OU Required",
                "Please select a Target OU before creating the group.\n\n"
                "Click '↻ Refresh OUs' to load available Organisational Units.")
            return
        ou = ou_dn

        if self.ad.connected and hasattr(self.ad, "conn") and self.ad.conn:
            ok, msg = self.ad.create_group(ou, name, gtype, desc, member_dns)
            if ok:
                self._result_lbl.config(fg=GREEN, text="✓ " + msg)
                self.log.log(f"Group created: {name} | Type: {gtype} | Members: {len(member_dns)}", "ok")
                self.status.set(f"Group '{name}' created successfully.", "ok")
            else:
                self._result_lbl.config(fg=RED, text="✗ " + msg)
                self.log.log(f"Group creation FAILED: {msg}", "error")
                self.status.set("Group creation failed — see log.", "error")
        else:
            # Demo
            m_str = ", ".join(member_lbls) if member_lbls else "None"
            msg = (f"[DEMO] Group prepared:\n"
                   f"  Name        : {name}\n"
                   f"  Type        : {gtype}\n"
                   f"  Description : {desc or '—'}\n"
                   f"  Members     : {m_str}\n"
                   f"  OU          : {ou},{self.ad.base_dn}")
            self._result_lbl.config(fg=TEAL, text="✓ Demo: Group would be created in AD")
            self.log.log(f"[DEMO] Group: {name} | Type: {gtype} | Members ({len(member_dns)}): {m_str}", "ok")
            self.status.set(f"[Demo] Group '{name}' prepared.", "ok")
            messagebox.showinfo("Demo — Group Created", msg)

    def _reset(self):
        self._name_var.set("")
        self._desc_var.set("")
        self._ou_dropdown.reset()
        self._type_var.set(self.GROUP_TYPES[0])
        self._result_lbl.config(text="")



# ─────────────────────────────────────────────────────────────────────────────
#  MANAGE GROUP FORM
# ─────────────────────────────────────────────────────────────────────────────
class ManageGroupForm(tk.Frame):
    def __init__(self, parent, ad: ADConnector, logger: ActivityLog, status: StatusBar):
        super().__init__(parent, bg=SURFACE)
        self.ad      = ad
        self.log     = logger
        self.status  = status
        self._groups = []
        self._filtered_groups = []
        self._sel_grp = None   # dict
        self._all_users = []
        self._build()

    def refresh_groups(self):
        if not self.ad.connected:
            return
        self._groups = self.ad.get_all_groups_full()
        self._filtered_groups = list(self._groups)
        self._populate_group_list()
        self.log.log(f"Manage Groups: {len(self._groups)} group(s) loaded.", "ok")

    def refresh_users(self):
        if self.ad.connected:
            raw = self.ad.get_all_users()
            self._all_users = [(f"{n} ({s})", dn) for n, s, dn, _ in raw]
            self._avail_lb_items = list(self._all_users)

    def _populate_group_list(self):
        self._glb.delete(0, "end")
        for g in self._filtered_groups:
            self._glb.insert("end", f"{g['cn']}  [{g['groupType']}]  ({g['memberCount']} members)")

    def _build(self):
        hdr = tk.Frame(self, bg="#F0FDF4", padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="👥  Manage Groups",
                 font=FONT_HEAD, fg=NAVY, bg="#F0FDF4").pack(side="left")
        tk.Label(hdr, text="Edit · Delete groups",
                 font=FONT_TINY, fg=TEXT2, bg="#F0FDF4").pack(side="right")

        paned = tk.PanedWindow(self, orient="horizontal", bg=BG,
                               sashwidth=6, relief="flat")
        paned.pack(fill="both", expand=True)

        # ── LEFT: group list ──────────────────────────────────────
        left = tk.Frame(paned, bg=SURFACE, width=310)
        paned.add(left, minsize=260)

        srch_row = tk.Frame(left, bg=SURFACE)
        srch_row.pack(fill="x", padx=8, pady=6)
        self._gsrch = tk.StringVar()
        self._gsrch.trace_add("write", self._filter_groups)
        tk.Entry(srch_row, textvariable=self._gsrch, font=FONT_BODY,
                 relief="solid", bd=1, bg=SURFACE2).pack(side="left", fill="x", expand=True)
        tk.Button(srch_row, text="↻", font=FONT_BODY, bg=LIGHT_BLU, fg=ACCENT,
                  relief="flat", padx=6, cursor="hand2",
                  command=self.refresh_groups).pack(side="right", padx=(4,0))

        lf = tk.Frame(left, bg=SURFACE)
        lf.pack(fill="both", expand=True, padx=8, pady=(0,8))
        self._glb = tk.Listbox(lf, font=FONT_BODY, bd=1, relief="solid",
                               selectbackground=TEAL, selectforeground=WHITE,
                               activestyle="none", bg=SURFACE2, fg=TEXT,
                               exportselection=False)
        gsb = ttk.Scrollbar(lf, orient="vertical", command=self._glb.yview)
        self._glb.configure(yscrollcommand=gsb.set)
        gsb.pack(side="right", fill="y")
        self._glb.pack(side="left", fill="both", expand=True)
        self._glb.bind("<<ListboxSelect>>", self._on_group_select)

        tk.Button(left, text="🗑  Delete Selected Group", font=FONT_TINY,
                  bg=RED, fg=WHITE, relief="flat", padx=8, pady=5,
                  cursor="hand2", command=self._delete_group).pack(
                  fill="x", padx=8, pady=(0,8))

        # ── RIGHT: edit form ──────────────────────────────────────
        right = tk.Frame(paned, bg=SURFACE)
        paned.add(right, minsize=400)

        self._no_sel_lbl = tk.Label(right,
            text="← Select a group from the list to edit",
            font=FONT_BODY, fg=TEXT3, bg=SURFACE)
        self._no_sel_lbl.place(relx=0.5, rely=0.4, anchor="center")

        self._edit_frame = tk.Frame(right, bg=SURFACE)

        canvas = tk.Canvas(self._edit_frame, bg=SURFACE, highlightthickness=0)
        esb = ttk.Scrollbar(self._edit_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=esb.set)
        esb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=SURFACE)
        win_id = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        _bind_mousewheel(canvas)

        pad = {"padx":12, "pady":3, "fill":"x"}
        SectionLabel(inner, "Group Details").pack(**pad)
        self._gname_var, _ = labeled_entry(inner, "Group Name  *", required=True)
        self._gdesc_var, _ = labeled_entry(inner, "Description", width=44)

        SectionLabel(inner, "Current Members").pack(**pad)
        mf = tk.Frame(inner, bg=SURFACE, padx=12)
        mf.pack(fill="x", pady=4)
        self._mem_lb = tk.Listbox(mf, font=FONT_BODY, height=6, bd=1,
                                  relief="solid", exportselection=False,
                                  selectbackground=RED, selectforeground=WHITE,
                                  activestyle="none", bg=SURFACE2, fg=TEXT,
                                  selectmode="multiple")
        msb = ttk.Scrollbar(mf, orient="vertical", command=self._mem_lb.yview)
        self._mem_lb.configure(yscrollcommand=msb.set)
        msb.pack(side="right", fill="y")
        self._mem_lb.pack(side="left", fill="both", expand=True)
        tk.Label(inner, text="Select members above then click Remove Selected",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE).pack(anchor="w", padx=12)
        tk.Button(inner, text="➖ Remove Selected Members", font=FONT_TINY,
                  bg=RED, fg=WHITE, relief="flat", padx=8, pady=4, cursor="hand2",
                  command=self._remove_members).pack(anchor="w", padx=12, pady=(2,6))

        SectionLabel(inner, "Add Members").pack(**pad)
        af = tk.Frame(inner, bg=SURFACE, padx=12)
        af.pack(fill="x", pady=4)
        self._add_lb = MultiSelectListbox(af, items=[], label="Search & select users to add:", height=5)
        self._add_lb.pack(fill="x")

        btn_row = tk.Frame(inner, bg=SURFACE, padx=12, pady=10)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="💾  Save Changes", font=FONT_SUB,
                  bg=TEAL, fg=WHITE, relief="flat", padx=16, pady=7,
                  cursor="hand2", command=self._save_group).pack(side="left", padx=(0,8))
        self._gedit_result = tk.Label(inner, text="", font=FONT_BODY,
                                      fg=GREEN, bg=SURFACE, wraplength=500)
        self._gedit_result.pack(fill="x", padx=12)

        self._cur_members = []   # [(label, dn)] of current members

    def _filter_groups(self, *_):
        q = self._gsrch.get().strip().lower()
        self._filtered_groups = [g for g in self._groups
                                  if q in g["cn"].lower()] if q else list(self._groups)
        self._populate_group_list()

    def _on_group_select(self, event=None):
        sel = self._glb.curselection()
        if not sel or sel[0] >= len(self._filtered_groups):
            return
        self._sel_grp = self._filtered_groups[sel[0]]
        self._no_sel_lbl.place_forget()
        self._edit_frame.place(relx=0, rely=0, relwidth=1, relheight=1)
        # populate fields
        self._gname_var.set(self._sel_grp["cn"])
        self._gdesc_var.set(self._sel_grp["description"])
        # load members
        if self.ad.connected and self.ad.conn:
            self._cur_members = self.ad.get_group_members(self._sel_grp["dn"])
        else:
            self._cur_members = []
        self._mem_lb.delete(0, "end")
        for label, _ in self._cur_members:
            self._mem_lb.insert("end", label)
        # populate add list
        self._add_lb.update_items(self._all_users)
        self._gedit_result.config(text="")

    def _remove_members(self):
        sel = self._mem_lb.curselection()
        if not sel:
            messagebox.showinfo("Select Members", "Select members to remove first.")
            return
        self._to_remove = [self._cur_members[i][1] for i in sel if i < len(self._cur_members)]
        for i in reversed(list(sel)):
            self._mem_lb.delete(i)
            if i < len(self._cur_members):
                self._cur_members.pop(i)

    def _save_group(self):
        if not self._sel_grp:
            return
        if not self.ad.connected or not getattr(self.ad.conn, "bound", False):
            messagebox.showwarning("Not Connected", "Connect to AD first.")
            return
        name = self._gname_var.get().strip()
        desc = self._gdesc_var.get().strip()
        old_dn     = self._sel_grp["dn"]
        add_dns    = self._add_lb.get_selected_values()
        remove_dns = getattr(self, "_to_remove", [])
        ok, msg = self.ad.modify_group(old_dn, name, desc, add_dns, remove_dns)
        if ok:
            _record("modif", "group", name)
            self._gedit_result.config(fg=GREEN, text="✓ " + msg)
            self.log.log(f"Group modified: {name}", "ok")
            self.status.set(f"Group '{name}' updated.", "ok")
            self._to_remove = []
            # If the name changed the DN changed too — update sel_grp so subsequent
            # operations (e.g. a second Save) reference the correct new DN
            old_cn = old_dn.split(",")[0].replace("CN=","").replace("cn=","").strip()
            if name and name != old_cn:
                parent = ",".join(old_dn.split(",")[1:])
                self._sel_grp = dict(self._sel_grp)   # make a mutable copy
                self._sel_grp["dn"] = f"CN={name},{parent}"
                self._sel_grp["cn"] = name
            self.after(600, self.refresh_groups)
        else:
            self._gedit_result.config(fg=RED, text="✗ " + msg)
            self.log.log(f"Group modify FAILED: {msg}", "error")

    def _delete_group(self):
        sel = self._glb.curselection()
        if not sel:
            messagebox.showwarning("Select Group", "Select a group to delete first.")
            return
        if not self.ad.connected or not getattr(self.ad.conn, "bound", False):
            messagebox.showwarning("Not Connected", "Connect to AD first.")
            return
        g = self._filtered_groups[sel[0]]
        if not messagebox.askyesno("Confirm Delete",
                f"Permanently delete group:\n\n  {g['cn']}\n\n"
                f"This cannot be undone.", icon="warning"):
            return
        ok, msg = self.ad.delete_group(g["dn"])
        if ok:
            _record("delet", "group", g["cn"])
            self.log.log(f"Group DELETED: {g['cn']}", "warn")
            self.status.set(f"Group '{g['cn']}' deleted.", "warn")
            self._sel_grp = None
            self._edit_frame.place_forget()
            self._no_sel_lbl.place(relx=0.5, rely=0.4, anchor="center")
            self.after(400, self.refresh_groups)
        else:
            messagebox.showerror("Delete Failed", msg)
            self.log.log(f"Group delete FAILED: {msg}", "error")


# ─────────────────────────────────────────────────────────────────────────────
#  REPORTS PANEL
# ─────────────────────────────────────────────────────────────────────────────
class ReportsPanel(tk.Frame):
    def __init__(self, parent, ad: ADConnector, logger: ActivityLog, status: StatusBar):
        super().__init__(parent, bg=SURFACE)
        self.ad      = ad
        self.log     = logger
        self.status  = status
        self._user_data  = []
        self._group_data = []
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg="#F5F3FF", padx=12, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📊  Reports",
                 font=FONT_HEAD, fg=NAVY, bg="#F5F3FF").pack(side="left")
        tk.Label(hdr, text="List · Export Users & Groups to XLSX / PDF",
                 font=FONT_TINY, fg=TEXT2, bg="#F5F3FF").pack(side="right")

        # Tab bar
        tab_bar = tk.Frame(self, bg=LIGHT_BLU)
        tab_bar.pack(fill="x")
        self._tab_btns = {}
        for label, key in [("👤 Users", "users"), ("👥 Groups", "groups"),
                            ("📋 Audit Log", "audit")]:
            b = tk.Button(tab_bar, text=label, font=FONT_BODY,
                          bg=ACCENT, fg=WHITE, relief="flat",
                          padx=16, pady=7, cursor="hand2",
                          command=lambda k=key: self._show_tab(k))
            b.pack(side="left", padx=2, pady=4)
            self._tab_btns[key] = b

        # Controls bar
        ctrl = tk.Frame(self, bg=SURFACE, padx=12, pady=6)
        ctrl.pack(fill="x")
        tk.Button(ctrl, text="↻ Refresh Data", font=FONT_TINY,
                  bg=LIGHT_BLU, fg=ACCENT, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self.refresh_all).pack(side="left", padx=(0,8))
        tk.Button(ctrl, text="📥 Export XLSX", font=FONT_TINY,
                  bg="#166534", fg=WHITE, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self._export_xlsx).pack(side="left", padx=(0,8))
        tk.Button(ctrl, text="📄 Export PDF", font=FONT_TINY,
                  bg="#7C3AED", fg=WHITE, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self._export_pdf).pack(side="left", padx=(0,8))
        self._rec_lbl = tk.Label(ctrl, text="", font=FONT_TINY, fg=TEXT2, bg=SURFACE)
        self._rec_lbl.pack(side="right")

        # Search
        srch_row = tk.Frame(self, bg=SURFACE, padx=12, pady=4)
        srch_row.pack(fill="x")
        self._srch_var = tk.StringVar()
        self._srch_var.trace_add("write", self._apply_filter)
        tk.Label(srch_row, text="🔍 Filter:", font=FONT_BODY,
                 fg=TEXT2, bg=SURFACE).pack(side="left")
        tk.Entry(srch_row, textvariable=self._srch_var, font=FONT_BODY,
                 relief="solid", bd=1, width=30, bg=SURFACE2).pack(side="left", padx=(6,0))

        # Treeview
        tree_frame = tk.Frame(self, bg=SURFACE)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=4)
        self._tree = ttk.Treeview(tree_frame, show="headings", selectmode="browse")
        tsb_y = ttk.Scrollbar(tree_frame, orient="vertical",   command=self._tree.yview)
        tsb_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=tsb_y.set, xscrollcommand=tsb_x.set)
        tsb_y.pack(side="right", fill="y")
        tsb_x.pack(side="bottom", fill="x")
        self._tree.pack(side="left", fill="both", expand=True)

        style = ttk.Style()
        style.configure("Treeview", rowheight=22, font=FONT_BODY)
        style.configure("Treeview.Heading", font=FONT_SUB)
        self._tree.tag_configure("disabled", foreground="#EF4444")
        self._tree.tag_configure("enabled",  foreground="#166534")

        self._cur_tab = "users"
        self._show_tab("users")

    def _show_tab(self, key):
        self._cur_tab = key
        for k, b in self._tab_btns.items():
            b.config(bg=ACCENT if k==key else BORDER,
                     fg=WHITE  if k==key else TEXT)
        if key == "users":
            self._load_user_table(self._user_data)
        elif key == "groups":
            self._load_group_table(self._group_data)
        else:
            self._load_audit_table()

    def _load_user_table(self, data):
        cols = ["Name","Username","Email","Department","Title","Company","Phone","Status","Created"]
        self._tree.config(columns=cols)
        for c in cols:
            self._tree.heading(c, text=c, command=lambda _c=c: self._sort(_c))
            self._tree.column(c, width=110, minwidth=60, stretch=True)
        self._tree.column("Name", width=160)
        self._tree.column("Email", width=180)
        self._tree.delete(*self._tree.get_children())
        for row in data:
            tag = "disabled" if row.get("Status")=="Disabled" else "enabled"
            self._tree.insert("", "end",
                values=[row.get(c,"") for c in cols], tags=(tag,))
        self._rec_lbl.config(text=f"{len(data)} user(s)")

    def _load_group_table(self, data):
        cols = ["Name","Type","Description","Members","DN"]
        self._tree.config(columns=cols)
        for c in cols:
            self._tree.heading(c, text=c, command=lambda _c=c: self._sort(_c))
            self._tree.column(c, width=130, minwidth=60, stretch=True)
        self._tree.column("DN", width=300)
        self._tree.delete(*self._tree.get_children())
        for row in data:
            self._tree.insert("", "end", values=[
                row.get("cn",""), row.get("groupType",""),
                row.get("description",""), row.get("memberCount",0),
                row.get("dn","")])
        self._rec_lbl.config(text=f"{len(data)} group(s)")

    def _load_audit_table(self):
        cols = ["Timestamp","Action","Type","Name"]
        self._tree.config(columns=cols)
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=160, minwidth=60, stretch=True)
        self._tree.delete(*self._tree.get_children())
        stats = _load_stats()
        for entry in reversed(stats.get("audit_log",[])):
            self._tree.insert("", "end", values=[
                entry.get("ts",""), entry.get("action",""),
                entry.get("type",""), entry.get("name","")])
        self._rec_lbl.config(text=f"{len(stats.get('audit_log',[]))} events")

    def _sort(self, col):
        data = [(self._tree.set(c, col), c) for c in self._tree.get_children("")]
        data.sort(key=lambda x: x[0].lower())
        for i, (_, iid) in enumerate(data):
            self._tree.move(iid, "", i)

    def _apply_filter(self, *_):
        q = self._srch_var.get().strip().lower()
        if self._cur_tab == "users":
            filtered = [r for r in self._user_data
                        if any(q in str(v).lower() for v in r.values())] if q else self._user_data
            self._load_user_table(filtered)
        elif self._cur_tab == "groups":
            filtered = [r for r in self._group_data
                        if q in r["cn"].lower() or q in r["description"].lower()] if q else self._group_data
            self._load_group_table(filtered)

    def refresh_all(self):
        if not self.ad.connected:
            return
        self.status.set("Fetching report data from AD…", "info")
        self._user_data  = self.ad.get_users_full()
        self._group_data = self.ad.get_all_groups_full()
        self._show_tab(self._cur_tab)
        self.log.log(f"Reports refreshed: {len(self._user_data)} users, "
                     f"{len(self._group_data)} groups.", "ok")
        self.status.set("Report data loaded.", "ok")

    def _get_current_rows(self):
        """Return (columns, rows) for whichever tab is active."""
        if self._cur_tab == "users":
            cols = ["Name","Username","Email","Department","Title","Company","Phone","Status","Created","DN"]
            rows = [[r.get(c,"") for c in cols] for r in self._user_data]
            return cols, rows
        elif self._cur_tab == "groups":
            cols = ["Name","Type","Description","Members","DN"]
            rows = [[r.get("cn",""), r.get("groupType",""),
                     r.get("description",""), r.get("memberCount",0),
                     r.get("dn","")] for r in self._group_data]
            return cols, rows
        else:
            stats = _load_stats()
            cols  = ["Timestamp","Action","Type","Name"]
            rows  = [[e.get("ts",""), e.get("action",""),
                      e.get("type",""), e.get("name","")] for e in stats.get("audit_log",[])]
            return cols, rows

    def _export_xlsx(self):
        from tkinter import filedialog
        tab  = self._cur_tab.capitalize()
        path = filedialog.asksaveasfilename(
            title=f"Export {tab} to Excel",
            defaultextension=".xlsx",
            initialfile=f"{tab}_Report_{datetime.date.today()}.xlsx",
            filetypes=[("Excel Files","*.xlsx"),("All Files","*.*")])
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl is required for Excel export.\n\n"
                "Install it: pip install openpyxl")
            return
        cols, rows = self._get_current_rows()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tab

        # Header styling
        hdr_fill = PatternFill("solid", fgColor="0F2D5E")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        thin     = Side(style="thin", color="CBD5E1")
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        # Data rows
        for ri, row in enumerate(rows, 2):
            fill = PatternFill("solid", fgColor="F8FAFC") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=str(val))
                cell.fill   = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        # Auto column widths
        for col_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 50)

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        stats = _load_stats()
        summary = [
            ("Report Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("AD Domain", self.ad.base_dn),
            ("",""),
            ("Total Users",   len(self._user_data)),
            ("Total Groups",  len(self._group_data)),
            ("",""),
            ("Users Created",  stats.get("users_created",0)),
            ("Users Modified", stats.get("users_modified",0)),
            ("Users Deleted",  stats.get("users_deleted",0)),
            ("Groups Created", stats.get("groups_created",0)),
            ("Groups Modified",stats.get("groups_modified",0)),
            ("Groups Deleted", stats.get("groups_deleted",0)),
        ]
        for r, (k,v) in enumerate(summary, 1):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=str(v))

        wb.save(path)
        self.log.log(f"XLSX exported: {os.path.basename(path)} ({len(rows)} rows)", "ok")
        self.status.set(f"XLSX saved: {os.path.basename(path)}", "ok")
        messagebox.showinfo("Export Complete",
            f"Excel report saved:\n{path}\n\n{len(rows)} records exported.")

    def _export_pdf(self):
        from tkinter import filedialog
        tab  = self._cur_tab.capitalize()
        path = filedialog.asksaveasfilename(
            title=f"Export {tab} to PDF",
            defaultextension=".pdf",
            initialfile=f"{tab}_Report_{datetime.date.today()}.pdf",
            filetypes=[("PDF Files","*.pdf"),("All Files","*.*")])
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            messagebox.showerror("Missing Library",
                "reportlab is required for PDF export.\n\n"
                "Install it: pip install reportlab")
            return

        cols, rows = self._get_current_rows()

        # ── Page setup ─────────────────────────────────────────────
        PAGE     = landscape(A4)
        L_MARGIN = 1.5 * cm
        R_MARGIN = 1.5 * cm
        doc = SimpleDocTemplate(path, pagesize=PAGE,
                                leftMargin=L_MARGIN, rightMargin=R_MARGIN,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        # ── Cell paragraph styles ─────────────────────────────────
        # Data cells — small font, word-wrap enabled
        cell_style = ParagraphStyle(
            "cell", parent=styles["Normal"],
            fontSize=6.5, leading=9,
            spaceAfter=0, spaceBefore=0,
            wordWrap="LTR",
        )
        # Header cells — white bold text
        hdr_style = ParagraphStyle(
            "hdr", parent=styles["Normal"],
            fontSize=7.5, leading=10,
            textColor=colors.white,
            fontName="Helvetica-Bold",
            wordWrap="LTR",
        )

        # ── Smart proportional column widths ──────────────────────
        # Weights reflect expected content length per column name.
        # Higher weight → wider column.
        WEIGHT_MAP = {
            # Users report
            "Name": 3.0, "Username": 2.0, "Email": 3.5,
            "Department": 2.5, "Title": 2.5, "Company": 2.5,
            "Phone": 1.8, "Status": 1.2, "Created": 1.5, "DN": 5.0,
            # Groups report
            "Type": 1.8, "Description": 3.5, "Members": 1.0,
            # Audit log
            "Timestamp": 2.0, "Action": 1.5,
        }
        avail_w = PAGE[0] - L_MARGIN - R_MARGIN   # usable page width
        weights  = [WEIGHT_MAP.get(c, 2.0) for c in cols]
        total_w  = sum(weights)
        col_w    = [(w / total_w) * avail_w for w in weights]

        # ── Helper: safe XML-escaped Paragraph ────────────────────
        def _para(val, style):
            text = str(val) if val is not None else ""
            # Escape XML special chars so ReportLab doesn't choke
            text = (text.replace("&", "&amp;")
                        .replace("<", "&lt;")
                        .replace(">", "&gt;"))
            return Paragraph(text, style)

        # ── Build table data with wrapped Paragraphs ──────────────
        table_data = [[_para(c, hdr_style) for c in cols]]
        for row in rows:
            table_data.append([_para(v, cell_style) for v in row])

        # ── Colours ───────────────────────────────────────────────
        navy  = colors.HexColor("#0F2D5E")
        alt   = colors.HexColor("#F0F4F8")
        red_c = colors.HexColor("#FEE2E2")
        grid  = colors.HexColor("#CBD5E1")

        tbl = Table(table_data, colWidths=col_w, repeatRows=1,
                    hAlign="LEFT")

        style_cmds = [
            # Header row
            ("BACKGROUND",    (0, 0), (-1, 0),  navy),
            ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
            ("VALIGN",        (0, 0), (-1, 0),  "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, 0),  5),
            ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
            ("LEFTPADDING",   (0, 0), (-1, 0),  4),
            ("RIGHTPADDING",  (0, 0), (-1, 0),  4),
            # Data rows
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, alt]),
            ("VALIGN",        (0, 1), (-1, -1), "TOP"),
            ("TOPPADDING",    (0, 1), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("LEFTPADDING",   (0, 1), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 1), (-1, -1), 4),
            # Grid
            ("GRID",          (0, 0), (-1, -1), 0.4, grid),
            ("BOX",           (0, 0), (-1, -1), 0.6, navy),
        ]

        # Highlight disabled user rows in light red
        if self._cur_tab == "users":
            status_col = cols.index("Status") if "Status" in cols else -1
            for ri, row in enumerate(rows, 1):
                if status_col >= 0 and str(row[status_col]) == "Disabled":
                    style_cmds.append(("BACKGROUND", (0, ri), (-1, ri), red_c))

        tbl.setStyle(TableStyle(style_cmds))

        # ── Title + metadata ──────────────────────────────────────
        title_style = ParagraphStyle(
            "rpt_title", parent=styles["Title"],
            fontSize=14, textColor=colors.HexColor("#0F2D5E"))
        meta_style = ParagraphStyle(
            "rpt_meta", parent=styles["Normal"],
            fontSize=8, textColor=colors.HexColor("#475569"))

        story = [
            Paragraph(f"Identity Manager — {tab} Report", title_style),
            Spacer(1, 0.15*cm),
            Paragraph(
                f"Domain: {self.ad.base_dn} &nbsp;&nbsp;·&nbsp;&nbsp; "
                f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} "
                f"&nbsp;&nbsp;·&nbsp;&nbsp; Records: {len(rows)}",
                meta_style),
            Spacer(1, 0.35*cm),
            tbl,
        ]

        doc.build(story)
        self.log.log(f"PDF exported: {os.path.basename(path)} ({len(rows)} rows)", "ok")
        self.status.set(f"PDF saved: {os.path.basename(path)}", "ok")
        messagebox.showinfo("Export Complete",
            f"PDF report saved:\n{path}\n\n{len(rows)} records exported.")


# ─────────────────────────────────────────────────────────────────────────────
#  DASHBOARD / HOME PANEL
# ─────────────────────────────────────────────────────────────────────────────
class DashboardPanel(tk.Frame):
    def __init__(self, parent, ad: ADConnector, on_user, on_group, on_manage,
                 on_manage_group, on_reports, logger: ActivityLog):
        super().__init__(parent, bg=SURFACE)
        self.ad = ad
        self._build(on_user, on_group, on_manage, on_manage_group, on_reports, logger)

    def _build(self, on_user, on_group, on_manage, on_manage_group, on_reports, logger):
        # Scrollable canvas for full dashboard
        canvas = tk.Canvas(self, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        _bind_mousewheel(canvas)

        # Hero
        hero = tk.Frame(inner, bg=NAVY, padx=32, pady=24)
        hero.pack(fill="x")
        tk.Label(hero, text="🏢  Identity Management App",
                 font=("Arial", 20, "bold"), fg=WHITE, bg=NAVY).pack(anchor="w")
        tk.Label(hero, text="Active Directory User & Group Administration",
                 font=("Arial", 12), fg="#93C5FD", bg=NAVY).pack(anchor="w", pady=(4,0))
        self._conn_lbl = tk.Label(hero, text="", font=FONT_BODY, fg="#FBBF24", bg=NAVY)
        self._conn_lbl.pack(anchor="w", pady=(8,0))
        self.update_connection_status()

        # ── Account Security Summary (refreshed live on every dashboard visit) ──
        stats_outer = tk.Frame(inner, bg=BG, padx=24, pady=16)
        stats_outer.pack(fill="x")
        tk.Label(stats_outer, text="Account Security Summary", font=FONT_SUB,
                 fg=NAVY, bg=BG).pack(anchor="w", pady=(0,10))
        stats_row = tk.Frame(stats_outer, bg=BG)
        stats_row.pack(fill="x")

        self._stat_labels = {}
        stat_defs = [
            ("locked_accounts",     "🔒 Locked Accounts",         RED),
            ("privileged_accounts", "🛡 Privileged Accounts",      AMBER),
            ("pwd_never_expire",    "🔑 Password Never Expires",   "#7C3AED"),
            ("disabled_accounts",   "⛔ Disabled Accounts",        "#475569"),
        ]
        for key, label, color in stat_defs:
            card = tk.Frame(stats_row, bg=SURFACE, relief="solid", bd=1,
                            padx=12, pady=10)
            card.pack(side="left", fill="both", expand=True, padx=(0,8))
            tk.Frame(card, bg=color, height=4).pack(fill="x")
            val_lbl = tk.Label(card, text="—", font=("Arial", 22, "bold"),
                               fg=color, bg=SURFACE)
            val_lbl.pack(pady=(6,2))
            tk.Label(card, text=label, font=FONT_TINY, fg=TEXT2,
                     bg=SURFACE, wraplength=110, justify="center").pack()
            self._stat_labels[key] = val_lbl
        tk.Label(stats_outer,
                 text="Privileged includes: Domain Admins · Enterprise Admins · Schema Admins · Administrator",
                 font=FONT_TINY, fg=TEXT3, bg=BG, anchor="w").pack(anchor="w", pady=(6,0))

        # ── Live AD counts ────────────────────────────────────────
        live_outer = tk.Frame(inner, bg=BG, padx=24, pady=0)
        live_outer.pack(fill="x")
        tk.Label(live_outer, text="Live AD Status", font=FONT_SUB,
                 fg=NAVY, bg=BG).pack(anchor="w", pady=(0,8))
        live_row1 = tk.Frame(live_outer, bg=BG)
        live_row1.pack(fill="x", pady=(0,8))
        live_row2 = tk.Frame(live_outer, bg=BG)
        live_row2.pack(fill="x", pady=(0,8))
        self._live_user_lbl    = self._make_live_card(live_row1, "👤 Total Users in AD",       ACCENT)
        self._live_group_lbl   = self._make_live_card(live_row1, "👥 Total Groups in AD",      TEAL)
        self._live_dis_lbl     = self._make_live_card(live_row1, "⛔ Disabled Accounts",       RED)
        self._live_locked_lbl  = self._make_live_card(live_row1, "🔒 Locked Accounts",         AMBER)
        self._live_priv_lbl    = self._make_live_card(live_row2, "🛡 Privileged (Admin) Users", "#7C3AED")
        self._live_noexp_lbl   = self._make_live_card(live_row2, "🔑 Password Never Expires",  "#0369A1")
        self._live_dis2_lbl    = self._make_live_card(live_row2, "⚠ Accounts Expiring Soon",  "#92400E")

        self.refresh_stats()

    def _make_live_card(self, parent, label, color):
        card = tk.Frame(parent, bg=SURFACE, relief="solid", bd=1, padx=14, pady=10)
        card.pack(side="left", fill="both", expand=True, padx=(0,8))
        tk.Frame(card, bg=color, height=4).pack(fill="x")
        lbl = tk.Label(card, text="—", font=("Arial",18,"bold"), fg=color, bg=SURFACE)
        lbl.pack(pady=(6,2))
        tk.Label(card, text=label, font=FONT_TINY, fg=TEXT2, bg=SURFACE,
                 wraplength=100, justify="center").pack()
        return lbl

    def _make_card(self, parent, icon, title, desc, color, cmd):
        card = tk.Frame(parent, bg=SURFACE, relief="solid", bd=1)
        tk.Frame(card, bg=color, height=5).pack(fill="x")
        body = tk.Frame(card, bg=SURFACE, padx=14, pady=12)
        body.pack(fill="both", expand=True)
        tk.Label(body, text=icon, font=("Arial",22), bg=SURFACE).pack(anchor="w")
        tk.Label(body, text=title, font=FONT_SUB, fg=NAVY, bg=SURFACE).pack(anchor="w",pady=(4,0))
        tk.Label(body, text=desc, font=FONT_TINY, fg=TEXT2, bg=SURFACE,
                 wraplength=160, justify="left").pack(anchor="w",pady=(4,10))
        tk.Button(body, text="Open →", font=FONT_BODY, bg=color, fg=WHITE,
                  relief="flat", padx=10, pady=5, cursor="hand2",
                  command=cmd).pack(anchor="w")
        return card

    def refresh_stats(self):
        """Reset security summary cards to a loading indicator.
        Actual values are pushed in by IdentityApp._update_dashboard_live_counts
        which runs in a background thread every time the dashboard is shown."""
        for lbl in self._stat_labels.values():
            lbl.config(text="…")

    def update_live_counts(self, n_users=None, n_groups=None, n_disabled=None,
                           n_locked=None, n_privileged=None, n_noexpiry=None,
                           n_expiring_soon=None):
        if n_users    is not None: self._live_user_lbl.config(text=str(n_users))
        if n_groups   is not None: self._live_group_lbl.config(text=str(n_groups))
        if n_disabled is not None: self._live_dis_lbl.config(text=str(n_disabled))
        if n_locked   is not None: self._live_locked_lbl.config(text=str(n_locked))
        if n_privileged  is not None: self._live_priv_lbl.config(text=str(n_privileged))
        if n_noexpiry    is not None: self._live_noexp_lbl.config(text=str(n_noexpiry))
        if n_expiring_soon is not None: self._live_dis2_lbl.config(text=str(n_expiring_soon))
        # ── Also push into the four security summary cards at the top ──────
        if n_locked     is not None:
            self._stat_labels["locked_accounts"].config(text=str(n_locked))
        if n_privileged is not None:
            self._stat_labels["privileged_accounts"].config(text=str(n_privileged))
        if n_noexpiry   is not None:
            self._stat_labels["pwd_never_expire"].config(text=str(n_noexpiry))
        if n_disabled   is not None:
            self._stat_labels["disabled_accounts"].config(text=str(n_disabled))

    def update_connection_status(self):
        if self.ad.connected:
            mode = "LDAPS" if self.ad.use_ssl else "LDAP"
            self._conn_lbl.config(
                text=f"✓ Connected ({mode})  |  Base DN: {self.ad.base_dn}",
                fg="#10d98a")
        else:
            self._conn_lbl.config(text="✗ Not connected to Active Directory", fg="#FBBF24")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN APPLICATION
# ─────────────────────────────────────────────────────────────────────────────
class IdentityApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Identity Management App — Active Directory")
        self.geometry("1100x780")
        self.minsize(900, 640)
        self.configure(bg=BG)
        self.ad = ADConnector()
        self._build_ui()
        # ── Mandatory activation check ────────────────────────────
        # Show the activation gate BEFORE the login window.
        # If the software is not active (never activated or expired),
        # the gate cannot be dismissed — the only exit is Quit.
        cfg = _load_config()
        act_ok, _, _ = _get_activation_status(cfg)
        if act_ok:
            self.after(200, self._show_login)
        else:
            self.after(200, self._show_activation_gate)

    # ── UI skeleton ──────────────────────────────────────────────
    def _build_ui(self):
        self._build_menu()

        # Main split: left nav + right content
        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True)

        # Left nav
        nav = tk.Frame(main, bg=NAVY, width=200)
        nav.pack(side="left", fill="y")
        nav.pack_propagate(False)
        tk.Label(nav, text="ID Manager", font=("Arial", 13, "bold"),
                 fg=WHITE, bg=NAVY, pady=18, padx=16).pack(fill="x")
        tk.Frame(nav, bg="#1A3A70", height=1).pack(fill="x")

        self._nav_btns = {}
        nav_items = [
            ("🏠  Dashboard",       "dashboard"),
            ("👤  Create User",     "user"),
            ("🔑  Manage User",     "manage"),
            ("👥  Create Group",    "group"),
            ("⚙   Manage Groups",  "managegroup"),
            ("📊  Reports",         "reports"),
            ("📋  Activity Log",    "log"),
        ]
        for label, key in nav_items:
            btn = tk.Button(nav, text=label, font=FONT_BODY, fg="#CBD5E1",
                            bg=NAVY, relief="flat", anchor="w", padx=20, pady=10,
                            cursor="hand2", activebackground=ACCENT2,
                            activeforeground=WHITE,
                            command=lambda k=key: self._show_panel(k))
            btn.pack(fill="x")
            self._nav_btns[key] = btn

        tk.Frame(nav, bg="#1A3A70", height=1).pack(fill="x", pady=(16, 0))
        self._conn_status = tk.Label(nav, text="● Disconnected",
                                     font=FONT_TINY, fg="#FBBF24", bg=NAVY,
                                     padx=16, pady=8, anchor="w")
        self._conn_status.pack(fill="x")

        tk.Button(nav, text="🔗  Connect / Auth", font=FONT_TINY, fg="#93C5FD",
                  bg="#1A3A70", relief="flat", padx=16, pady=6, cursor="hand2",
                  command=self._show_login).pack(fill="x", padx=8, pady=4)

        # Spacer pushes logo+footer to bottom
        tk.Frame(nav, bg=NAVY).pack(fill="both", expand=True)

        # Logo area
        self._logo_lbl = tk.Label(nav, bg=NAVY, cursor="hand2")
        self._logo_lbl.pack(fill="x", pady=(0, 2))
        self._logo_lbl.bind("<Button-1>", lambda e: self._upload_logo())
        self._logo_placeholder = tk.Label(nav,
            text="[ Click to upload\n  company logo ]",
            font=FONT_TINY, fg="#475569", bg=NAVY, pady=6)
        self._logo_placeholder.pack(fill="x")
        self._logo_placeholder.bind("<Button-1>", lambda e: self._upload_logo())
        self._load_saved_logo()

        # Footer
        tk.Frame(nav, bg="#0A1C3B", height=1).pack(fill="x")

        # Right content area
        right = tk.Frame(main, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Content stack
        self._content = tk.Frame(right, bg=BG)
        self._content.pack(fill="both", expand=True)

        # Vendor footer — activation status only
        footer = tk.Frame(right, bg="#0A1C3B", height=22)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        cfg_f = _load_config()
        act_ok_f, act_rem_f, _ = _get_activation_status(cfg_f)
        act_ico  = "✓" if act_ok_f else "⚠"
        act_col  = "#10d98a" if act_ok_f else "#FBBF24"
        act_foot = f"{act_ico} {'Active · ' + str(act_rem_f) + ' days left' if act_ok_f else 'Not activated — Help > Activate'}"
        tk.Label(footer, text=act_foot, font=("Arial", 8),
                 fg=act_col, bg="#0A1C3B").pack(side="left", padx=12)
        tk.Label(footer, text=_VENDOR_NAME, font=("Arial", 8, "bold"),
                 fg="#3B82F6", bg="#0A1C3B", padx=12).pack(side="right")

        # Status bar (global - stays visible on all panels)
        self._status = StatusBar(right)
        self._status.pack(fill="x", side="bottom")

        # Logger — buffer-based, no visible widget here; output goes to Log tab only
        self._logger = ActivityLog()

        # Build panels
        self._panels = {}
        self._panels["dashboard"] = DashboardPanel(
            self._content, self.ad,
            on_user         = lambda: self._show_panel("user"),
            on_group        = lambda: self._show_panel("group"),
            on_manage       = lambda: self._show_panel("manage"),
            on_manage_group = lambda: self._show_panel("managegroup"),
            on_reports      = lambda: self._show_panel("reports"),
            logger          = self._logger
        )
        self._panels["user"]        = UserCreationForm(self._content, self.ad, self._logger, self._status)
        self._panels["manage"]      = ManageUserForm(self._content, self.ad, self._logger, self._status)
        self._panels["group"]       = GroupCreationForm(self._content, self.ad, self._logger, self._status)
        self._panels["managegroup"] = ManageGroupForm(self._content, self.ad, self._logger, self._status)
        self._panels["reports"]     = ReportsPanel(self._content, self.ad, self._logger, self._status)
        self._panels["log"]         = self._build_log_panel()

        for p in self._panels.values():
            p.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._show_panel("dashboard")

    def _build_log_panel(self):
        frame = tk.Frame(self._content, bg=SURFACE)
        hdr = tk.Frame(frame, bg=LIGHT_BLU, padx=16, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📋  Activity Log",
                 font=FONT_HEAD, fg=NAVY, bg=LIGHT_BLU).pack(side="left")
        big = scrolledtext.ScrolledText(frame, font=FONT_MONO,
                                        bg="#0F172A", fg="#94A3B8",
                                        relief="flat", bd=0, state="disabled")
        big.pack(fill="both", expand=True, padx=8, pady=8)
        tk.Button(frame, text="Clear Log", font=FONT_TINY,
                  bg=BORDER, fg=TEXT, relief="flat", padx=8, pady=4,
                  cursor="hand2",
                  command=lambda: [big.config(state="normal"),
                                   big.delete("1.0", "end"),
                                   big.config(state="disabled")]).pack(
                  anchor="e", padx=8, pady=(0, 8))
        # Register this ScrolledText with the buffer logger
        self._logger.register_widget(big)
        return frame

    def _build_menu(self):
        mbar = tk.Menu(self)
        self.config(menu=mbar)
        # File
        file_m = tk.Menu(mbar, tearoff=0)
        mbar.add_cascade(label="File", menu=file_m)
        file_m.add_command(label="Connect to AD…", command=self._show_login)
        file_m.add_command(label="Disconnect", command=self._disconnect)
        file_m.add_separator()
        file_m.add_command(label="📷 Upload Company Logo…", command=self._upload_logo)
        file_m.add_separator()
        file_m.add_command(label="Exit", command=self.quit)
        # Actions
        act_m = tk.Menu(mbar, tearoff=0)
        mbar.add_cascade(label="Actions", menu=act_m)
        act_m.add_command(label="Create User Account",                    command=lambda: self._show_panel("user"))
        act_m.add_command(label="Manage User (Password/Enable/Edit/Delete)", command=lambda: self._show_panel("manage"))
        act_m.add_command(label="Create Group",                            command=lambda: self._show_panel("group"))
        act_m.add_command(label="Manage Groups (Edit/Delete)",             command=lambda: self._show_panel("managegroup"))
        act_m.add_command(label="Reports",                                 command=lambda: self._show_panel("reports"))
        act_m.add_separator()
        act_m.add_command(label="Refresh OUs (for User form)",
                          command=lambda: self._panels["user"].refresh_ous())
        act_m.add_command(label="Refresh Groups (for User form)",
                          command=lambda: self._panels["user"].refresh_groups())
        act_m.add_command(label="Refresh OUs (for Group form)",
                          command=lambda: self._panels["group"].refresh_ous())
        act_m.add_command(label="Refresh Users (for Group form)",
                          command=lambda: self._panels["group"].refresh_users())
        # Help
        help_m = tk.Menu(mbar, tearoff=0)
        mbar.add_cascade(label="Help", menu=help_m)
        help_m.add_command(label="🔑 Activate Software…", command=self._show_activate)
        help_m.add_separator()
        help_m.add_command(label="About", command=self._about)

    # ── Panel switching ──────────────────────────────────────────
    def _show_panel(self, key):
        for k, p in self._panels.items():
            p.lift() if k == key else p.lower()
        for k, b in self._nav_btns.items():
            b.config(bg=ACCENT if k == key else NAVY,
                     fg=WHITE if k == key else "#CBD5E1")
        if key == "dashboard":
            self._panels["dashboard"].update_connection_status()
            self._panels["dashboard"].refresh_stats()
            # Refresh live AD counts in background so the dashboard always shows
            # up-to-date figures when the user returns to it
            if self.ad.connected:
                self.after(100, self._update_dashboard_live_counts)

    def _update_dashboard_live_counts(self):
        """Fetch live AD metrics in a background thread and push results
        to the dashboard labels on the main thread via after(0, …).
        This prevents the UI from freezing while LDAP queries run."""
        if not self.ad.connected:
            return

        def _fetch():
            try:
                users  = self.ad.get_users_full()
                groups = self.ad.get_all_groups_full()
                n_dis  = sum(1 for u in users if u.get("Status") == "Disabled")
                n_locked     = 0
                n_privileged = 0
                n_noexpiry   = 0
                n_expiring   = 0
                try:
                    from ldap3 import SUBTREE
                    import datetime as _dt
                    # Single LDAP search for all per-user security attributes
                    self.ad.conn.search(
                        search_base   = self.ad.base_dn,
                        search_filter = "(&(objectClass=user)(objectCategory=person))",
                        search_scope  = SUBTREE,
                        attributes    = ["userAccountControl", "lockoutTime",
                                         "memberOf", "accountExpires"],
                        size_limit    = 2000
                    )
                    priv_keywords = ("domain admins", "enterprise admins",
                                     "schema admins", "group policy creator",
                                     "cn=administrator,")
                    today     = _dt.date.today()
                    soon_days = 30
                    for e in self.ad.conn.entries:
                        try:
                            uac  = int(str(e.userAccountControl)) if e.userAccountControl else 512
                            lock = int(str(e.lockoutTime))        if e.lockoutTime        else 0
                            if lock > 0:
                                n_locked += 1
                            if uac & 0x10000:           # DONT_EXPIRE_PASSWORD
                                n_noexpiry += 1
                            member_of = e.memberOf.values if hasattr(e.memberOf, "values") else []
                            for g in member_of:
                                if any(kw in str(g).lower() for kw in priv_keywords):
                                    n_privileged += 1
                                    break
                            acc_raw = e.accountExpires
                            if acc_raw:
                                acc_val = int(str(acc_raw))
                                if acc_val not in (0, 9223372036854775807):
                                    epoch_diff = 116444736000000000
                                    exp_ts   = (acc_val - epoch_diff) / 10_000_000
                                    exp_date = _dt.date.fromtimestamp(exp_ts)
                                    if 0 <= (exp_date - today).days <= soon_days:
                                        n_expiring += 1
                        except Exception:
                            pass
                except Exception:
                    pass
                # Marshal results back to the main (Tk) thread
                self.after(0, lambda: self._panels["dashboard"].update_live_counts(
                    n_users        = len(users),
                    n_groups       = len(groups),
                    n_disabled     = n_dis,
                    n_locked       = n_locked,
                    n_privileged   = n_privileged,
                    n_noexpiry     = n_noexpiry,
                    n_expiring_soon= n_expiring,
                ))
            except Exception:
                pass

        threading.Thread(target=_fetch, daemon=True).start()

    # ── Login / connection ───────────────────────────────────────

    def _load_saved_logo(self):
        """Load logo from disk if previously saved."""
        try:
            from PIL import Image, ImageTk
            raw = _load_logo_data()
            if raw:
                img = Image.open(__import__("io").BytesIO(raw))
                img.thumbnail((160, 80), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                self._logo_lbl.config(image=self._logo_img)
                self._logo_placeholder.pack_forget()
        except Exception:
            pass  # PIL not installed or no logo — show placeholder

    def _upload_logo(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select Company Logo",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.gif *.ico"),
                       ("All files", "*.*")]
        )
        if not path:
            return
        try:
            from PIL import Image, ImageTk
            img = Image.open(path)
            raw = open(path, "rb").read()
            _save_logo_data(raw)
            img.thumbnail((160, 80), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            self._logo_lbl.config(image=self._logo_img)
            self._logo_placeholder.pack_forget()
            self._logger.log(f"Company logo saved: {os.path.basename(path)}", "ok")
        except ImportError:
            # PIL not available — show filename as text fallback
            self._logo_placeholder.config(
                text=f"Logo: {os.path.basename(path)}\n(install Pillow for preview)",
                fg="#93C5FD")
            _save_logo_data(open(path, "rb").read())
            self._logger.log("Logo saved (install Pillow for visual preview): pip install Pillow", "warn")
        except Exception as ex:
            messagebox.showerror("Logo Error", f"Could not load logo:\n{ex}")

    def _show_activate(self):
        """Activation dialog — enter code to unlock for N months."""
        dlg = tk.Toplevel(self)
        dlg.title("Activate Identity Manager")
        dlg.resizable(False, False)
        dlg.configure(bg=SURFACE)
        dlg.grab_set()
        dlg.update_idletasks()

        cfg = _load_config()
        act_ok, act_rem, act_exp = _get_activation_status(cfg)

        # Header
        hdr = tk.Frame(dlg, bg=NAVY)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔑  Activate Identity Manager",
                 font=FONT_HEAD, fg=WHITE, bg=NAVY, pady=12, padx=16).pack(side="left")

        body = tk.Frame(dlg, bg=SURFACE, padx=24, pady=16)
        body.pack(fill="both", expand=True)

        # Current status
        status_col = "#10d98a" if act_ok else "#FBBF24"
        status_txt = (f"✓ ACTIVE — {act_rem} days remaining  (expires {act_exp})"
                      if act_ok else f"⚠ {act_exp}")
        tk.Label(body, text="Current status:", font=FONT_BODY, fg=TEXT2,
                 bg=SURFACE, anchor="w").pack(fill="x")
        tk.Label(body, text=status_txt, font=("Arial", 10, "bold"),
                 fg=status_col, bg=SURFACE, anchor="w",
                 pady=4).pack(fill="x")

        tk.Frame(body, bg=BORDER, height=1).pack(fill="x", pady=10)

        # Activation code entry
        tk.Label(body, text="Enter Activation Code:", font=FONT_BODY,
                 fg=TEXT2, bg=SURFACE, anchor="w").pack(fill="x")
        code_var = tk.StringVar()
        code_ent = tk.Entry(body, textvariable=code_var, font=("Courier New", 13),
                            width=28, relief="solid", bd=2, justify="center",
                            highlightthickness=1, highlightcolor=ACCENT)
        code_ent.pack(pady=(4, 2))
        code_ent.focus()

        info_lbl = tk.Label(body, text="Format: IDP-DAYS-XXXXXXXXXXXXXXXX-YYYYYYYY",
                            font=FONT_TINY, fg=TEXT3, bg=SURFACE)
        info_lbl.pack()

        result_lbl = tk.Label(body, text="", font=FONT_BODY, fg=GREEN,
                              bg=SURFACE, wraplength=380)
        result_lbl.pack(pady=8)

        def _activate():
            code = code_var.get().strip().upper()
            days = _verify_activation_code(code)
            if days <= 0:
                result_lbl.config(fg=RED, text="✗ Invalid activation code. Please check and try again.")
                return

            cfg2 = _load_config()

            # ── Reuse prevention via SHA-256 hash ─────────────────────────
            # Store hashes, not plain codes, so deleting from config doesn't help.
            code_hash   = hashlib.sha256(code.encode()).hexdigest()
            used_hashes = cfg2.get("used_code_hashes", [])
            used_plain  = cfg2.get("used_codes", [])   # legacy compat
            if code_hash in used_hashes or code in used_plain:
                result_lbl.config(fg=RED,
                    text="✗ This activation code has already been used on this machine "
                         "and cannot be reused.\n"
                         "Each code is single-use. Contact ITProAcademy.co.in for a new code.")
                return

            # ── First use — activate ──────────────────────────────────────
            used_hashes.append(code_hash)
            used_plain.append(code)
            cfg2["used_code_hashes"] = used_hashes
            cfg2["used_codes"]       = used_plain
            cfg2["activated_on"]     = datetime.date.today().isoformat()
            cfg2["activated_days"]   = days
            cfg2["activation_code"]  = code
            _save_config(cfg2)

            expiry = datetime.date.today() + datetime.timedelta(days=days)
            months = round(days / 30, 1)
            result_lbl.config(fg=GREEN,
                text=f"✓ Activated for {days} days ({months} months).\n"
                     f"Expires: {expiry.isoformat()}\n"
                     f"This code is now marked as used and cannot be reused.")

        tk.Frame(body, bg=BORDER, height=1).pack(fill="x", pady=(4,10))
        btn_row = tk.Frame(body, bg=SURFACE)
        btn_row.pack(fill="x")
        tk.Button(btn_row, text="✓  Activate", font=FONT_SUB, bg=ACCENT,
                  fg=WHITE, relief="flat", padx=20, pady=6, cursor="hand2",
                  command=_activate).pack(side="left", padx=(0,8))
        tk.Button(btn_row, text="Close", font=FONT_BODY, bg=BORDER,
                  fg=TEXT, relief="flat", padx=12, pady=6, cursor="hand2",
                  command=dlg.destroy).pack(side="left")

        tk.Frame(body, bg=BORDER, height=1).pack(fill="x", pady=(12,6))
        tk.Label(body,
            text=f"Contact {_VENDOR_NAME} to obtain your activation code. Each code activates the software for a fixed number of days.",
            font=FONT_TINY, fg=TEXT3, bg=SURFACE, justify="left").pack(anchor="w")

        # Center dialog
        dlg.update_idletasks()
        w, h = dlg.winfo_width(), dlg.winfo_height()
        x = (dlg.winfo_screenwidth()  - w) // 2
        y = (dlg.winfo_screenheight() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

    # ── Mandatory activation gate (blocks the app until activated) ──
    def _show_activation_gate(self):
        """
        Full-screen modal that cannot be dismissed.
        The user must enter a valid, never-before-used activation code
        or click Quit to exit the application.
        """
        gate = tk.Toplevel(self)
        gate.title("Activation Required — Identity Manager")
        gate.resizable(False, False)
        gate.configure(bg=SURFACE)
        # Make it truly modal and uncloseable via the X button
        gate.grab_set()
        gate.protocol("WM_DELETE_WINDOW", lambda: self.quit())
        gate.lift()

        # ── Header ────────────────────────────────────────────────
        hdr = tk.Frame(gate, bg=NAVY)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔐  Activation Required",
                 font=("Arial", 16, "bold"), fg=WHITE, bg=NAVY,
                 pady=16, padx=20).pack(side="left")
        tk.Label(hdr, text=f"v{_APP_VERSION}",
                 font=FONT_TINY, fg="#93C5FD", bg=NAVY,
                 padx=16).pack(side="right")

        # ── Warning banner ────────────────────────────────────────
        warn = tk.Frame(gate, bg="#7C1A1A", padx=20, pady=10)
        warn.pack(fill="x")
        cfg = _load_config()
        _, _, act_exp = _get_activation_status(cfg)
        expired = cfg.get("activated_on") and not _get_activation_status(cfg)[0]
        if expired:
            warn_txt = f"⚠  Your licence has expired ({act_exp}). Enter a new activation code to continue."
        else:
            warn_txt = "⚠  This software requires a valid activation code before it can be used."
        tk.Label(warn, text=warn_txt,
                 font=("Arial", 9, "bold"), fg="#FEE2E2", bg="#7C1A1A",
                 wraplength=440, justify="left").pack(anchor="w")

        # ── Body ──────────────────────────────────────────────────
        body = tk.Frame(gate, bg=SURFACE, padx=28, pady=20)
        body.pack(fill="both", expand=True)

        tk.Label(body,
                 text="Please enter your activation code to unlock Identity Manager.\n"
                      "Contact ITProAcademy.co.in to obtain a code.",
                 font=FONT_BODY, fg=TEXT2, bg=SURFACE,
                 justify="left").pack(anchor="w", pady=(0, 14))

        code_row = tk.Frame(body, bg=SURFACE)
        code_row.pack(fill="x")
        tk.Label(code_row, text="Activation Code:", font=FONT_BODY,
                 fg=TEXT, bg=SURFACE, width=16, anchor="w").pack(side="left")
        gate_code_var = tk.StringVar()
        gate_code_ent = tk.Entry(code_row, textvariable=gate_code_var,
                                 font=("Courier New", 13), width=26,
                                 relief="solid", bd=2, justify="center",
                                 highlightthickness=2, highlightcolor=ACCENT, bg=SURFACE2)
        gate_code_ent.pack(side="left", padx=(6, 0))
        gate_code_ent.focus()

        tk.Label(body, text="Format: IDP-DAYS-XXXXXXXXXXXXXXXX-YYYYYYYY",
                 font=FONT_TINY, fg=TEXT3, bg=SURFACE).pack(anchor="w", pady=(4, 0))

        gate_result = tk.Label(body, text="", font=FONT_BODY,
                               fg=RED, bg=SURFACE, wraplength=440)
        gate_result.pack(fill="x", pady=10)

        def _do_gate_activate():
            code = gate_code_var.get().strip().upper()
            if not code:
                gate_result.config(fg=RED, text="✗ Please enter an activation code.")
                return
            days = _verify_activation_code(code)
            if days <= 0:
                gate_result.config(fg=RED,
                    text="✗ Invalid activation code. Check the code and try again.\n"
                         "Contact ITProAcademy.co.in if you need a new code.")
                return

            cfg2 = _load_config()
            # ── Reuse check using SHA-256 hash of the code ────────
            used_hashes = cfg2.get("used_code_hashes", [])
            code_hash = hashlib.sha256(code.encode()).hexdigest()
            if code_hash in used_hashes:
                gate_result.config(fg=RED,
                    text="✗ This activation code has already been used on this machine "
                         "and cannot be reused.\n"
                         "Contact ITProAcademy.co.in for a new code.")
                return

            # ── Activate ──────────────────────────────────────────
            used_hashes.append(code_hash)
            cfg2["used_code_hashes"] = used_hashes
            # Keep legacy plain list in sync for _show_activate dialog
            used_plain = cfg2.get("used_codes", [])
            used_plain.append(code)
            cfg2["used_codes"]      = used_plain
            cfg2["activated_on"]    = datetime.date.today().isoformat()
            cfg2["activated_days"]  = days
            cfg2["activation_code"] = code
            _save_config(cfg2)

            expiry = datetime.date.today() + datetime.timedelta(days=days)
            gate_result.config(fg=GREEN,
                text=f"✓ Activated for {days} days.  Expires: {expiry.isoformat()}")
            gate.update()
            gate.after(900, lambda: [gate.destroy(), self._show_login()])

        btn_row = tk.Frame(body, bg=SURFACE)
        btn_row.pack(fill="x", pady=(4, 0))
        act_btn = tk.Button(btn_row, text="✓  Activate & Continue",
                            font=FONT_SUB, bg=ACCENT, fg=WHITE,
                            relief="flat", padx=20, pady=8, cursor="hand2",
                            command=_do_gate_activate)
        act_btn.pack(side="left", padx=(0, 10))
        act_btn.bind("<Enter>", lambda e: act_btn.config(bg=ACCENT2))
        act_btn.bind("<Leave>", lambda e: act_btn.config(bg=ACCENT))
        gate_code_ent.bind("<Return>", lambda e: _do_gate_activate())

        tk.Button(btn_row, text="✕  Quit", font=FONT_BODY,
                  bg="#94A3B8", fg=WHITE, relief="flat", padx=12, pady=8,
                  cursor="hand2", command=self.quit).pack(side="left")

        # ── Footer ────────────────────────────────────────────────
        tk.Frame(gate, bg=BORDER, height=1).pack(fill="x")
        ft = tk.Frame(gate, bg="#F8FAFC", padx=20, pady=8)
        ft.pack(fill="x")
        tk.Label(ft, text=f"Powered by {_VENDOR_NAME}  ·  Each activation code is single-use.",
                 font=FONT_TINY, fg=TEXT3, bg="#F8FAFC").pack(side="left")

        # Centre the gate over the main window
        gate.update_idletasks()
        gw, gh = gate.winfo_reqwidth(), gate.winfo_reqheight()
        sx = self.winfo_x() + (self.winfo_width()  - gw) // 2
        sy = self.winfo_y() + (self.winfo_height() - gh) // 2
        gate.geometry(f"{gw}x{gh}+{sx}+{sy}")

    def _show_login(self):
        LoginWindow(self, self.ad, self._on_connected)

    def _on_connected(self, real=True):
        if real:
            mode = "LDAPS" if self.ad.use_ssl else "LDAP"
            self._conn_status.config(
                text=f"● Connected ({mode})", fg="#10d98a")
            self._logger.log(
                f"Connected to AD: {self.ad.base_dn} | "
                f"{'LDAPS SSL' if self.ad.use_ssl else 'Plain LDAP — StartTLS for passwords'}",
                "ok")
            self._status.set("Connected to AD. Loading OUs, groups, users…", "ok")
            # Update dashboard immediately — don't wait for tab switch
            self._panels["dashboard"].update_connection_status()
            self._panels["dashboard"].refresh_stats()
            # Refresh the password warning now we know the SSL state
            self.after(200, self._panels["user"]._update_pwd_warning)
            # Sequential deferred refreshes — staggered to avoid overwhelming AD
            self.after(400,  self._panels["user"].refresh_ous)
            self.after(900,  self._panels["user"].refresh_groups)
            self.after(1400, self._panels["user"].refresh_users_for_manager)
            self.after(1900, self._panels["group"].refresh_ous)
            self.after(2000, self._panels["group"].refresh_users)
            self.after(2100, self._panels["manage"].refresh_users)
            self.after(2200, self._panels["managegroup"].refresh_groups)
            self.after(2300, self._panels["managegroup"].refresh_users)
            self.after(2400, self._panels["reports"].refresh_all)
            self.after(2600, self._update_dashboard_live_counts)
            self.after(2700, lambda: self._status.set(
                "AD data loaded. Navigate to Create User, Manage User, or Create Group.", "ok"))
        else:
            self._conn_status.config(text="● Demo Mode", fg="#FBBF24")
            # Load demo OUs
            demo_ous = [
                ("[Root — DC=demo,DC=local]",   "DC=demo,DC=local"),
                ("Users",                        "OU=Users,DC=demo,DC=local"),
                ("Computers",                    "OU=Computers,DC=demo,DC=local"),
                ("Groups",                       "OU=Groups,DC=demo,DC=local"),
                ("Service Accounts",             "OU=Service Accounts,DC=demo,DC=local"),
                ("Users > IT",                   "OU=IT,OU=Users,DC=demo,DC=local"),
                ("Users > HR",                   "OU=HR,OU=Users,DC=demo,DC=local"),
                ("Users > Finance",              "OU=Finance,OU=Users,DC=demo,DC=local"),
                ("Users > Management",           "OU=Management,OU=Users,DC=demo,DC=local"),
            ]
            # Load demo groups and users
            demo_groups = [
                ("IT-Admins",         "CN=IT-Admins,OU=Groups,DC=demo,DC=local"),
                ("HR-Department",     "CN=HR-Department,OU=Groups,DC=demo,DC=local"),
                ("Finance-Team",      "CN=Finance-Team,OU=Groups,DC=demo,DC=local"),
                ("All-Staff",         "CN=All-Staff,OU=Groups,DC=demo,DC=local"),
                ("DevOps-Engineers",  "CN=DevOps-Engineers,OU=Groups,DC=demo,DC=local"),
                ("Security-Team",     "CN=Security-Team,OU=Groups,DC=demo,DC=local"),
                ("Management",        "CN=Management,OU=Groups,DC=demo,DC=local"),
                ("Remote-Users",      "CN=Remote-Users,OU=Groups,DC=demo,DC=local"),
            ]
            demo_users = [
                ("Alice Johnson <alice@demo.local>",  "CN=Alice Johnson,OU=Users,DC=demo,DC=local"),
                ("Bob Smith <bob@demo.local>",        "CN=Bob Smith,OU=Users,DC=demo,DC=local"),
                ("Carol Williams <carol@demo.local>", "CN=Carol Williams,OU=Users,DC=demo,DC=local"),
                ("David Brown <david@demo.local>",    "CN=David Brown,OU=Users,DC=demo,DC=local"),
                ("Eve Davis <eve@demo.local>",        "CN=Eve Davis,OU=Users,DC=demo,DC=local"),
                ("Frank Miller <frank@demo.local>",   "CN=Frank Miller,OU=Users,DC=demo,DC=local"),
                ("Grace Wilson <grace@demo.local>",   "CN=Grace Wilson,OU=Users,DC=demo,DC=local"),
                ("Henry Moore <henry@demo.local>",    "CN=Henry Moore,OU=Users,DC=demo,DC=local"),
            ]
            self._panels["user"]._ous = demo_ous
            self._panels["user"]._ou_dropdown.update_ous(demo_ous)
            self._panels["user"]._groups = demo_groups
            self._panels["user"]._grp_selector.update_items(demo_groups)
            self._panels["group"]._ous = demo_ous
            self._panels["group"]._ou_dropdown.update_ous(demo_ous)
            self._panels["group"]._users = demo_users
            self._panels["group"]._user_selector.update_items(demo_users)
            # Manage panel — populate with enriched demo users (name, sam, dn, mail, uac, lock)
            demo_manage = [
                ("Alice Johnson",  "alice",  "CN=Alice Johnson,OU=Users,DC=demo,DC=local",  "alice@demo.local",  512, 0),
                ("Bob Smith",      "bob",    "CN=Bob Smith,OU=Users,DC=demo,DC=local",      "bob@demo.local",    514, 0),
                ("Carol Williams", "carol",  "CN=Carol Williams,OU=Users,DC=demo,DC=local", "carol@demo.local",  512, 0),
                ("David Brown",    "david",  "CN=David Brown,OU=Users,DC=demo,DC=local",    "david@demo.local",  514, 1000000),
                ("Eve Davis",      "eve",    "CN=Eve Davis,OU=Users,DC=demo,DC=local",      "eve@demo.local",    512, 0),
                ("Frank Miller",   "frank",  "CN=Frank Miller,OU=Users,DC=demo,DC=local",   "frank@demo.local",  512, 0),
                ("Grace Wilson",   "grace",  "CN=Grace Wilson,OU=Users,DC=demo,DC=local",   "grace@demo.local",  514, 0),
                ("Henry Moore",    "henry",  "CN=Henry Moore,OU=Users,DC=demo,DC=local",    "henry@demo.local",  512, 0),
            ]
            self._panels["manage"]._users = demo_manage
            self._panels["manage"]._filtered = demo_manage
            self._panels["manage"]._populate_list(demo_manage)
            self._logger.log("Running in Demo Mode — no real AD connection", "warn")
            self._status.set("Demo Mode active — simulated AD data loaded.", "warn")
        # Always update dashboard immediately regardless of mode
        self._panels["dashboard"].update_connection_status()
        self._panels["dashboard"].refresh_stats()
        self._show_panel("dashboard")

    def _disconnect(self):
        self.ad.disconnect()
        self._conn_status.config(text="● Disconnected", fg="#FBBF24")
        self._logger.log("Disconnected from AD.", "warn")
        self._status.set("Disconnected.", "warn")
        self._panels["dashboard"].update_connection_status()

    def _about(self):
        cfg = _load_config()
        act_ok, act_rem, act_exp = _get_activation_status(cfg)
        act_str = (f"ACTIVE — {act_rem} days remaining (expires {act_exp})"
                   if act_ok else f"NOT ACTIVATED ({act_exp})")
        messagebox.showinfo("About",
            "Identity Management App v1.0\n\n"
            "Active Directory User & Group Administration\n\n"
            "Built with: Python 3 · tkinter · ldap3\n\n"
            "Supports:\n"
            "  • User Account Creation (all AD attributes)\n"
            "  • Group Creation (Security & Distribution)\n"
            "  • Real AD integration via LDAP/LDAPS\n"
            "  • Demo Mode for offline demonstration\n\n"
            "Requirements: Python 3.8+ · pip install ldap3\n\n"
            f"Activation: {act_str}\n\n"
            f"Powered by: {_VENDOR_NAME}\n"
            "For real AD usage: ensure LDAP port 389 (or LDAPS 636)\n"
            "is reachable from this machine to the domain controller.")


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = IdentityApp()
    app.mainloop()
